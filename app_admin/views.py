from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count, F, Q
from django.utils import timezone
from django.views.decorators.http import require_POST
from datetime import date, timedelta
import json
import hmac
import hashlib
from django.conf import settings

from authentication.models import Perfil, Estudiante, Padre, Docente
from .models import (
    Producto, Categoria, Ingrediente, RecetaIngrediente,
    LoteIngrediente, ProduccionElaborado,
    Proveedor, CompraProveedor, DetalleCompra,
    Pedido, DetallePedido,
    MovimientoInventario, MovimientoIngrediente,
    Insumo, MovimientoInsumo,
    PerfilAdmin, Alergeno,
    GoogleCalendarToken,
    descontar_fifo, verificar_lotes_vencidos,
)
from .forms import LoteIngredienteForm, ProduccionForm
from app_docente.models import PedidoDocente, DetallePedidoDocente


# ══════════════════════════════════════════════════════════════════════════════
# DECORADOR
# ══════════════════════════════════════════════════════════════════════════════

def admin_required(view_func):
    @login_required(login_url='/login/')
    def wrapper(request, *args, **kwargs):
        try:
            if request.user.perfil.rol != 'admin':
                messages.error(request, 'No tienes permisos de administrador.')
                return redirect('authentication:login')
        except Perfil.DoesNotExist:
            return redirect('authentication:login')
        return view_func(request, *args, **kwargs)
    wrapper.__name__ = view_func.__name__
    return wrapper


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _get_alertas():
    """Devuelve conteo de alertas activas para el topbar."""
    hoy = date.today()
    stock_prod = Producto.objects.filter(stock__gt=0, stock__lte=F('stock_minimo')).count()

    # Stock de ingredientes: calcula con base en lotes (property stock_real)
    todos_ings = list(Ingrediente.objects.filter(activo=True).prefetch_related('lotes'))
    stock_ing  = sum(1 for i in todos_ings if i.stock_bajo)

    # Lotes con stock positivo que ya vencieron
    vencidos = LoteIngrediente.objects.filter(
        fecha_vencimiento__lt=hoy, cantidad_base__gt=0
    ).values('ingrediente').distinct().count()

    # Lotes que vencen en ≤ 7 días
    vence_pronto = LoteIngrediente.objects.filter(
        cantidad_base__gt=0,
        fecha_vencimiento__gte=hoy,
        fecha_vencimiento__lte=hoy + timedelta(days=7)
    ).values('ingrediente').distinct().count()

    pedidos_pend = Pedido.objects.filter(
        estado__in=['pendiente', 'preparando'],
        fecha_pedido__date=hoy
    ).count()
    pedidos_pend += PedidoDocente.objects.filter(
        estado__in=['pendiente', 'preparando'],
        fecha_pedido__date=hoy
    ).count()

    total = stock_prod + stock_ing + vencidos + vence_pronto + pedidos_pend
    return {
        'total':        total,
        'stock_prod':   stock_prod,
        'stock_ing':    stock_ing,
        'vencidos':     vencidos,
        'vence_pronto': vence_pronto,
        'pedidos_pend': pedidos_pend,
    }


def _ical_token():
    """Token estable derivado de SECRET_KEY para autenticar el feed iCal."""
    return hmac.new(
        settings.SECRET_KEY.encode(),
        b'ical_feed_puntoasis',
        hashlib.sha256,
    ).hexdigest()[:40]


def _ctx(extra=None):
    """Contexto base con alertas para todas las vistas admin."""
    ctx = {'alertas': _get_alertas()}
    if extra:
        ctx.update(extra)
    return ctx


# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════

@admin_required
def dashboard(request):
    hoy = date.today()
    # Solo contar pedidos entregados para ingresos reales
    pedidos_hoy  = Pedido.objects.filter(fecha_pedido__date=hoy, estado='entregado')

    # Un solo agregado para ambos valores
    agg = pedidos_hoy.aggregate(ingresos=Sum('total'), costos=Sum('costo_total'))
    ventas_hoy   = float(agg['ingresos'] or 0)
    costos_hoy   = float(agg['costos'] or 0)
    ganancia_hoy = ventas_hoy - costos_hoy
    pendientes  = pedidos_hoy.filter(estado__in=['pendiente', 'preparando']).count()
    entregados  = pedidos_hoy.filter(estado='entregado').count()

    # Stock crítico: productos E ingredientes juntos
    prods_criticos = Producto.objects.filter(
        tipo='simple', stock__lte=F('stock_minimo')
    ).select_related('proveedor').order_by('stock')[:8]

    todos_ings_act = list(Ingrediente.objects.filter(activo=True).prefetch_related('lotes'))
    ings_criticos = sorted(
        [i for i in todos_ings_act if i.stock_bajo or i.sin_stock],
        key=lambda i: i.stock_real
    )[:8]

    ultimos_pedidos = Pedido.objects.select_related(
        'estudiante__perfil__user'
    ).order_by('-fecha_pedido')[:8]

    context = _ctx({
        'ventas_hoy':      ventas_hoy,
        'ganancia_hoy':    ganancia_hoy,
        'pendientes':      pendientes,
        'entregados':      entregados,
        'total_pedidos':   pedidos_hoy.count(),
        'prods_criticos':  prods_criticos,
        'ings_criticos':   ings_criticos,
        'ultimos_pedidos': ultimos_pedidos,
        'fecha_hoy':       hoy,
    })
    return render(request, 'app_admin/dashboard.html', context)


# ══════════════════════════════════════════════════════════════════════════════
# PRODUCTOS
# ══════════════════════════════════════════════════════════════════════════════

@admin_required
def productos(request):
    qs = Producto.objects.select_related('categoria', 'proveedor').prefetch_related('receta__ingrediente').all()
    cat_f  = request.GET.get('categoria', '')
    tipo_f = request.GET.get('tipo', '')
    disp_f = request.GET.get('disponible', '')
    q      = request.GET.get('q', '').strip()

    if cat_f:  qs = qs.filter(categoria__nombre=cat_f)
    if tipo_f: qs = qs.filter(tipo=tipo_f)
    if disp_f == '1': qs = qs.filter(disponible=True)
    if disp_f == '0': qs = qs.filter(disponible=False)
    if q: qs = qs.filter(nombre__icontains=q)

    lista = list(qs)
    kpi = {
        'total':      len(lista),
        'sin_stock':  sum(1 for p in lista if p.sin_stock),
        'stock_bajo': sum(1 for p in lista if p.stock_bajo and not p.sin_stock),
        'elaborados': sum(1 for p in lista if p.tipo == 'elaborado'),
    }
    movimientos = MovimientoInventario.objects.select_related('producto').order_by('-fecha')[:20]

    return render(request, 'app_admin/productos.html', _ctx({
        'productos':   qs,
        'categorias':  Categoria.objects.filter(activa=True),
        'cat_f': cat_f, 'tipo_f': tipo_f, 'disp_f': disp_f, 'q': q,
        'kpi':         kpi,
        'movimientos': movimientos,
    }))


@admin_required
def producto_nuevo(request):
    categorias   = Categoria.objects.filter(activa=True)
    proveedores  = Proveedor.objects.filter(activo=True)
    ingredientes = Ingrediente.objects.filter(activo=True)

    if request.method == 'POST':
        tipo     = request.POST.get('tipo', 'simple')
        nombre   = request.POST.get('nombre', '').strip()
        desc     = request.POST.get('descripcion', '').strip()
        precio_v = request.POST.get('precio_venta')
        precio_c = request.POST.get('precio_costo') or None
        cat_id   = request.POST.get('categoria')
        prov_id  = request.POST.get('proveedor') or None
        stock    = request.POST.get('stock', 0)
        stock_min = request.POST.get('stock_minimo', 5)
        disponible = request.POST.get('disponible') == 'on'
        imagen   = request.FILES.get('imagen')

        errores = []
        if not nombre or not precio_v or not cat_id:
            errores.append('Nombre, precio de venta y categoría son obligatorios.')
        if tipo == 'simple' and not prov_id:
            errores.append('Los productos simples requieren un proveedor.')

        if errores:
            for e in errores:
                messages.error(request, e)
        else:
            p = Producto.objects.create(
                tipo=tipo, nombre=nombre, descripcion=desc,
                precio_venta=precio_v,
                precio_costo=precio_c if tipo == 'simple' else None,
                categoria_id=cat_id,
                proveedor_id=prov_id if tipo == 'simple' else None,
                stock=stock if tipo == 'simple' else 0,
                stock_minimo=stock_min,
                disponible=disponible, imagen=imagen,
            )
            if tipo == 'elaborado':
                for i_id, i_cant in zip(
                    request.POST.getlist('ingrediente_id'),
                    request.POST.getlist('ingrediente_cant')
                ):
                    if i_id and i_cant:
                        RecetaIngrediente.objects.create(producto=p, ingrediente_id=i_id, cantidad=i_cant)
            messages.success(request, f'Producto "{p.nombre}" creado.')
            return redirect('app_admin:productos')

    return render(request, 'app_admin/producto_form.html', _ctx({
        'categorias': categorias, 'proveedores': proveedores,
        'ingredientes': ingredientes, 'accion': 'Nuevo producto',
    }))


@admin_required
def producto_editar(request, pk):
    producto     = get_object_or_404(Producto, pk=pk)
    categorias   = Categoria.objects.filter(activa=True)
    proveedores  = Proveedor.objects.filter(activo=True)
    ingredientes = Ingrediente.objects.filter(activo=True)
    receta       = producto.receta.select_related('ingrediente').all()

    if request.method == 'POST':
        tipo_orig = producto.tipo
        producto.nombre       = request.POST.get('nombre', producto.nombre).strip()
        producto.descripcion  = request.POST.get('descripcion', '').strip()
        producto.precio_venta = request.POST.get('precio_venta', producto.precio_venta)
        producto.disponible   = request.POST.get('disponible') == 'on'
        cat_id = request.POST.get('categoria')
        if cat_id:
            producto.categoria_id = cat_id

        if producto.tipo == 'simple':
            prov_id = request.POST.get('proveedor') or None
            if not prov_id:
                messages.error(request, 'Los productos simples requieren un proveedor.')
                return render(request, 'app_admin/producto_form.html', _ctx({
                    'producto': producto, 'categorias': categorias,
                    'proveedores': proveedores, 'ingredientes': ingredientes,
                    'receta': receta, 'accion': 'Editar producto',
                }))
            producto.precio_costo = request.POST.get('precio_costo') or None
            producto.proveedor_id = prov_id
            producto.stock        = request.POST.get('stock', producto.stock)
            producto.stock_minimo = request.POST.get('stock_minimo', producto.stock_minimo)

        if request.FILES.get('imagen'):
            producto.imagen = request.FILES['imagen']
        producto.save()

        if producto.tipo == 'elaborado':
            with transaction.atomic():
                producto.receta.all().delete()
                for i_id, i_cant in zip(
                    request.POST.getlist('ingrediente_id'),
                    request.POST.getlist('ingrediente_cant')
                ):
                    if i_id and i_cant:
                        RecetaIngrediente.objects.create(producto=producto, ingrediente_id=i_id, cantidad=i_cant)

        messages.success(request, f'Producto "{producto.nombre}" actualizado.')
        return redirect('app_admin:productos')

    return render(request, 'app_admin/producto_form.html', _ctx({
        'producto': producto, 'categorias': categorias,
        'proveedores': proveedores, 'ingredientes': ingredientes,
        'receta': receta, 'accion': 'Editar producto',
    }))


@admin_required
def producto_eliminar(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        nombre = producto.nombre
        producto.delete()
        messages.success(request, f'Producto "{nombre}" eliminado.')
    return redirect('app_admin:productos')


@admin_required
def producto_toggle(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        producto.disponible = not producto.disponible
        producto.save(update_fields=['disponible'])
    return redirect('app_admin:productos')


@admin_required
def producto_bulk(request):
    """Acciones masivas sobre productos: ocultar, mostrar o eliminar."""
    if request.method == 'POST':
        pks    = request.POST.getlist('pks')
        accion = request.POST.get('accion', '')
        if not pks:
            messages.warning(request, 'No se seleccionó ningún producto.')
            return redirect('app_admin:productos')
        productos_qs = Producto.objects.filter(pk__in=pks)
        count = productos_qs.count()
        if accion == 'ocultar':
            productos_qs.update(disponible=False)
            messages.success(request, f'{count} producto(s) ocultado(s) del menú.')
        elif accion == 'mostrar':
            productos_qs.update(disponible=True)
            messages.success(request, f'{count} producto(s) visible(s) en el menú.')
        elif accion == 'eliminar':
            productos_qs.delete()
            messages.success(request, f'{count} producto(s) eliminado(s).')
        else:
            messages.error(request, 'Acción inválida.')
    return redirect('app_admin:productos')


@admin_required
def ingrediente_bulk(request):
    """Acciones masivas sobre ingredientes: desactivar."""
    if request.method == 'POST':
        pks    = request.POST.getlist('pks')
        accion = request.POST.get('accion', '')
        if not pks:
            messages.warning(request, 'No se seleccionó ningún ingrediente.')
            return redirect('app_admin:ingredientes')
        qs = Ingrediente.objects.filter(pk__in=pks)
        count = qs.count()
        if accion == 'eliminar':
            qs.update(activo=False)
            messages.success(request, f'{count} ingrediente(s) desactivado(s).')
        else:
            messages.error(request, 'Acción inválida.')
    return redirect('app_admin:ingredientes')


@admin_required
def insumo_bulk(request):
    """Acciones masivas sobre insumos: desactivar."""
    if request.method == 'POST':
        pks    = request.POST.getlist('pks')
        accion = request.POST.get('accion', '')
        if not pks:
            messages.warning(request, 'No se seleccionó ningún insumo.')
            return redirect('app_admin:insumos')
        qs = Insumo.objects.filter(pk__in=pks)
        count = qs.count()
        if accion == 'eliminar':
            qs.update(activo=False)
            messages.success(request, f'{count} insumo(s) desactivado(s).')
        else:
            messages.error(request, 'Acción inválida.')
    return redirect('app_admin:insumos')


# ══════════════════════════════════════════════════════════════════════════════
# INVENTARIO
# ══════════════════════════════════════════════════════════════════════════════

@admin_required
def inventario(request):
    qs = Producto.objects.select_related('categoria', 'proveedor').all()
    cat_f  = request.GET.get('categoria', '')
    tipo_f = request.GET.get('tipo', '')
    q      = request.GET.get('q', '').strip()
    if cat_f:  qs = qs.filter(categoria__nombre=cat_f)
    if tipo_f: qs = qs.filter(tipo=tipo_f)
    if q:      qs = qs.filter(Q(nombre__icontains=q) | Q(proveedor__nombre__icontains=q))

    ings = Ingrediente.objects.filter(activo=True)
    if q: ings = ings.filter(nombre__icontains=q)

    insumos_qs = Insumo.objects.select_related('proveedor').filter(activo=True)
    if q: insumos_qs = insumos_qs.filter(Q(nombre__icontains=q) | Q(proveedor__nombre__icontains=q))

    movimientos = MovimientoInventario.objects.select_related(
        'producto', 'compra__proveedor'
    ).order_by('-fecha')[:30]

    movimientos_ings = MovimientoIngrediente.objects.select_related(
        'ingrediente', 'compra__proveedor'
    ).order_by('-fecha')[:30]

    return render(request, 'app_admin/inventario.html', _ctx({
        'productos':        qs,
        'ingredientes':     ings,
        'insumos':          insumos_qs,
        'movimientos':      movimientos,
        'movimientos_ings': movimientos_ings,
        'categorias':       Categoria.objects.filter(activa=True),
        'cat_f': cat_f, 'tipo_f': tipo_f, 'q': q,
    }))


@admin_required
def inventario_ajuste(request, pk):
    """Ajuste/salida/merma — las entradas van por entrada_stock."""
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        tipo     = request.POST.get('tipo', 'ajuste')
        cantidad = int(request.POST.get('cantidad', 0))
        nota     = request.POST.get('nota', '').strip()

        if tipo == 'salida':
            producto.stock = max(0, producto.stock - cantidad)
        elif tipo == 'ajuste':
            producto.stock = cantidad
        elif tipo == 'merma':
            producto.stock = max(0, producto.stock - cantidad)

        producto.save(update_fields=['stock'])
        MovimientoInventario.objects.create(
            producto=producto, tipo=tipo, cantidad=cantidad, nota=nota
        )
        messages.success(request, f'Inventario de "{producto.nombre}" actualizado.')
    return redirect('app_admin:inventario')


@admin_required
def inventario_historial(request, pk):
    producto    = get_object_or_404(Producto, pk=pk)
    movimientos = producto.movimientos.select_related('compra__proveedor').order_by('-fecha')
    return render(request, 'app_admin/inventario_historial.html', _ctx({
        'producto':    producto,
        'movimientos': movimientos,
    }))


# ──────────────────────────────────────────────────────────────────────────────
# ENTRADA UNIFICADA DE STOCK
# ──────────────────────────────────────────────────────────────────────────────

@admin_required
def entrada_stock(request, pk=None):
    """
    Formulario único de entrada de stock.
    Siempre crea una CompraProveedor y vincula el movimiento.
    Si se llama con pk, preselecciona ese producto.
    Si viene ?ing=<pk>, preselecciona ese ingrediente.
    """
    proveedores  = Proveedor.objects.filter(activo=True)
    productos    = Producto.objects.filter(tipo='simple').select_related('proveedor')
    ingredientes = Ingrediente.objects.filter(activo=True)
    producto_sel    = get_object_or_404(Producto, pk=pk) if pk else None
    ing_pk          = request.GET.get('ing', '').strip()
    ingrediente_sel = get_object_or_404(Ingrediente, pk=ing_pk) if ing_pk else None

    if request.method == 'POST':
        prov_id  = request.POST.get('proveedor')
        fecha    = request.POST.get('fecha', str(date.today()))
        nota     = request.POST.get('nota', '').strip()

        if not prov_id:
            messages.error(request, 'Debes seleccionar un proveedor.')
            return render(request, 'app_admin/entrada_stock.html', _ctx({
                'proveedores': proveedores, 'productos': productos,
                'ingredientes': ingredientes, 'producto_sel': producto_sel,
                'ingrediente_sel': ingrediente_sel,
            }))

        item_tipos   = request.POST.getlist('item_tipo')
        item_ids     = request.POST.getlist('item_id')
        item_cants   = request.POST.getlist('item_cantidad')
        item_precios = request.POST.getlist('item_precio')
        item_vencs   = request.POST.getlist('item_vencimiento')

        # Pre-validar todos los ítems antes de tocar la BD.
        # Conservamos el índice original para resolver el vencimiento posicional
        # incluso cuando el mismo ítem aparece varias veces (Bug #2c).
        lineas_validas = []
        for idx, (tipo, id_, cant, precio) in enumerate(zip(item_tipos, item_ids, item_cants, item_precios)):
            if not (id_ and cant and precio):
                continue
            if tipo not in ('producto', 'ingrediente'):
                messages.error(request, 'Tipo de ítem inválido.')
                return redirect('app_admin:entrada_stock')
            try:
                cant_f   = float(cant)
                precio_f = float(precio)
                if cant_f <= 0 or precio_f < 0:
                    raise ValueError
            except (ValueError, TypeError):
                messages.error(request, 'Cantidad y precio deben ser números positivos.')
                return redirect('app_admin:entrada_stock')
            # Bug #6/7: productos simples solo aceptan unidades enteras
            if tipo == 'producto' and cant_f != int(cant_f):
                messages.error(request, 'Los productos se compran en unidades enteras.')
                return redirect('app_admin:entrada_stock')
            venc = item_vencs[idx] if idx < len(item_vencs) else ''
            lineas_validas.append((tipo, id_, cant_f, precio_f, venc))

        if not lineas_validas:
            messages.error(request, 'No hay ítems válidos en la entrada.')
            return redirect('app_admin:entrada_stock')

        # Todo en una sola transacción atómica
        with transaction.atomic():
            compra = CompraProveedor.objects.create(
                proveedor_id=prov_id, fecha=fecha, nota=nota
            )
            total = 0
            for tipo, id_, cant_f, precio_f, venc in lineas_validas:
                det = DetalleCompra(compra=compra, cantidad=cant_f, precio_unitario=precio_f)
                if tipo == 'producto':
                    prod = get_object_or_404(Producto, pk=id_)
                    det.producto_id = prod.pk
                    prod.stock += int(cant_f)
                    prod.save(update_fields=['stock'])
                    MovimientoInventario.objects.create(
                        producto=prod, tipo='entrada',
                        cantidad=int(cant_f),
                        nota=nota or f'Compra #{compra.pk}',
                        compra=compra
                    )
                else:
                    ing = get_object_or_404(Ingrediente, pk=id_)
                    det.ingrediente_id = ing.pk
                    f_venc = venc.strip() if venc else None
                    if not f_venc:
                        from datetime import date as _date
                        f_venc = str(_date.today().replace(year=_date.today().year + 1))
                    cpd = float(ing.contenido_por_unidad) or 1
                    cantidad_base = round(cant_f * cpd, 3)
                    LoteIngrediente.objects.create(
                        ingrediente=ing,
                        proveedor_id=prov_id,
                        unidades_compra=cant_f,
                        precio_compra=precio_f,
                        cantidad_base=cantidad_base,
                        cantidad_base_inicial=cantidad_base,
                        fecha_vencimiento=f_venc,
                        compra=compra,
                        nota=nota or f'Compra #{compra.pk}',
                    )
                    MovimientoIngrediente.objects.create(
                        ingrediente=ing, tipo='entrada',
                        cantidad=cantidad_base,
                        nota=nota or f'Compra #{compra.pk} — lote vence {f_venc}',
                        compra=compra
                    )
                det.save()
                total += cant_f * precio_f

            compra.total = total
            compra.save(update_fields=['total'])

        messages.success(request, f'Entrada registrada. Compra #{compra.pk} — Total: ${total:,.0f}')
        return redirect('app_admin:proveedor_compras', pk=prov_id)

    # Auto-select proveedor desde el producto seleccionado
    prov_autosel = ''
    if producto_sel and producto_sel.proveedor:
        prov_autosel = str(producto_sel.proveedor.pk)

    return render(request, 'app_admin/entrada_stock.html', _ctx({
        'proveedores':     proveedores,
        'productos':       productos,
        'ingredientes':    ingredientes,
        'producto_sel':    producto_sel,
        'ingrediente_sel': ingrediente_sel,
        'prov_get_pk':     request.GET.get('prov', '') or prov_autosel,
    }))


@admin_required
def salida_stock(request):
    """Registrar salida/merma de stock para productos o ingredientes."""
    productos    = Producto.objects.filter(tipo='simple').select_related('proveedor')
    ingredientes = Ingrediente.objects.filter(activo=True)

    # Pre-selección via GET params: ?tipo=ingrediente&item=<pk>
    item_tipo_sel = request.GET.get('tipo', 'ingrediente')
    item_pk_sel   = request.GET.get('item', '')

    if request.method == 'POST':
        tipo_item = request.POST.get('item_tipo', 'ingrediente')
        item_id   = request.POST.get('item_id', '').strip()
        cantidad  = request.POST.get('cantidad', '0').strip()
        motivo    = request.POST.get('motivo', 'salida')
        nota      = request.POST.get('nota', '').strip()

        try:
            cantidad = float(cantidad)
        except ValueError:
            cantidad = 0

        if not item_id or cantidad <= 0:
            messages.error(request, 'Selecciona un ítem y especifica una cantidad mayor a 0.')
        else:
            try:
                if tipo_item == 'producto':
                    prod = get_object_or_404(Producto, pk=item_id)
                    # Bug #6/7: productos se manejan en unidades enteras
                    if cantidad != int(cantidad):
                        messages.error(request, f'Los productos se retiran en unidades enteras (ingresaste {cantidad}).')
                        return redirect('app_admin:salida_stock')
                    cantidad_int = int(cantidad)
                    # Bug #8: rechazar salida cuando excede el stock disponible
                    if cantidad_int > prod.stock:
                        messages.error(request, f'No puedes retirar {cantidad_int} unidades de "{prod.nombre}". Stock disponible: {prod.stock}.')
                        return redirect('app_admin:salida_stock')
                    prod.stock -= cantidad_int
                    prod.save(update_fields=['stock'])
                    MovimientoInventario.objects.create(
                        producto=prod, tipo=motivo,
                        cantidad=cantidad_int, nota=nota
                    )
                    messages.success(request, f'Salida de "{prod.nombre}" registrada. Stock actual: {prod.stock}')
                else:
                    ing = get_object_or_404(Ingrediente, pk=item_id)
                    # Bug #8: rechazar salida cuando excede stock real (suma de lotes)
                    disponible = float(ing.stock_real)
                    if cantidad > disponible:
                        messages.error(request, f'No puedes retirar {cantidad:.2f} {ing.get_unidad_base_display()} de "{ing.nombre}". Disponible: {disponible:.2f}.')
                        return redirect('app_admin:salida_stock')
                    with transaction.atomic():
                        resto = descontar_fifo(ing, cantidad, nota=nota)
                        descontado = cantidad - resto
                        MovimientoIngrediente.objects.create(
                            ingrediente=ing, tipo=motivo,
                            cantidad=descontado, nota=nota
                        )
                    messages.success(request, f'Salida de "{ing.nombre}" registrada. Stock real: {ing.stock_real:.2f} {ing.get_unidad_base_display()}')
                return redirect('app_admin:inventario')
            except Exception as e:
                messages.error(request, f'Error al registrar la salida: {str(e)}')

    productos_json = json.dumps([
        {'id': p.pk, 'nombre': p.nombre, 'stock': p.stock, 'unidad': 'und'}
        for p in productos
    ])
    ingredientes_json = json.dumps([
        {
            'id': ing.pk, 'nombre': ing.nombre,
            'stock': ing.stock_real,
            'unidad': ing.get_unidad_base_display(),
        }
        for ing in ingredientes
    ])

    return render(request, 'app_admin/salida_stock.html', _ctx({
        'productos_json':    productos_json,
        'ingredientes_json': ingredientes_json,
        'item_tipo_sel':     item_tipo_sel,
        'item_pk_sel':       item_pk_sel,
    }))


@admin_required
def ingrediente_eliminar(request, pk):
    ing = get_object_or_404(Ingrediente, pk=pk)
    if request.method == 'POST':
        nombre = ing.nombre
        ing.activo = False
        ing.save(update_fields=['activo'])
        messages.success(request, f'Ingrediente "{nombre}" desactivado del inventario.')
    return redirect('app_admin:inventario')


# ══════════════════════════════════════════════════════════════════════════════
# INGREDIENTES
# ══════════════════════════════════════════════════════════════════════════════

@admin_required
def ingredientes(request):
    # Auto-expire any lots that passed their date
    n_vencidos = verificar_lotes_vencidos()

    qs = Ingrediente.objects.select_related('proveedor').prefetch_related('alergenos', 'lotes').filter(activo=True)
    q  = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(nombre__icontains=q)

    movimientos = MovimientoIngrediente.objects.select_related(
        'ingrediente', 'compra__proveedor'
    ).order_by('-fecha')[:50]

    todos = list(Ingrediente.objects.filter(activo=True).prefetch_related('lotes'))
    kpi = {
        'sin_stock':      sum(1 for i in todos if i.sin_stock),
        'stock_bajo':     sum(1 for i in todos if i.stock_bajo),
        'vence_pronto':   sum(1 for i in todos if i.vence_pronto),
        'total':          len(todos),
        'n_vencidos_hoy': n_vencidos,
    }

    return render(request, 'app_admin/ingredientes.html', _ctx({
        'ingredientes': qs,
        'movimientos':  movimientos,
        'q':            q,
        'kpi':          kpi,
    }))


@admin_required
def ingrediente_nuevo(request):
    alergenos_todos = Alergeno.objects.all()
    proveedores     = Proveedor.objects.filter(activo=True)

    if request.method == 'POST':
        nombre               = request.POST.get('nombre', '').strip()
        proveedor_id         = request.POST.get('proveedor') or None
        unidad_compra        = request.POST.get('unidad_compra', 'und').strip()
        contenido_por_unidad = request.POST.get('contenido_por_unidad', 1)
        unidad_base          = request.POST.get('unidad_base', 'und')
        stock_min            = request.POST.get('stock_minimo', 0)
        imagen               = request.FILES.get('imagen')
        alerg_ids            = request.POST.getlist('alergenos')

        errores = []
        if not nombre:
            errores.append('El nombre es obligatorio.')
        if not proveedor_id:
            errores.append('Debes seleccionar el proveedor principal.')

        if errores:
            for e in errores:
                messages.error(request, e)
        else:
            ing = Ingrediente.objects.create(
                nombre=nombre,
                proveedor_id=proveedor_id,
                unidad_compra=unidad_compra,
                contenido_por_unidad=contenido_por_unidad,
                unidad_base=unidad_base,
                stock_minimo=stock_min,
                imagen=imagen,
            )
            if alerg_ids:
                ing.alergenos.set(alerg_ids)
            messages.success(request, f'Ingrediente "{nombre}" creado. Ahora registra el primer lote.')
            return redirect('app_admin:lotes_ingrediente', pk=ing.pk)

    return render(request, 'app_admin/ingrediente_form.html', _ctx({
        'alergenos_todos': alergenos_todos,
        'proveedores':     proveedores,
        'accion':          'Nuevo ingrediente',
    }))


@admin_required
def ingrediente_editar(request, pk):
    ing             = get_object_or_404(Ingrediente, pk=pk)
    alergenos_todos = Alergeno.objects.all()
    proveedores     = Proveedor.objects.filter(activo=True)

    if request.method == 'POST':
        ing.nombre    = request.POST.get('nombre', ing.nombre).strip()
        prov_id       = request.POST.get('proveedor') or None
        if not prov_id:
            messages.error(request, 'El proveedor principal es obligatorio.')
            return render(request, 'app_admin/ingrediente_form.html', _ctx({
                'ingrediente': ing, 'alergenos_todos': alergenos_todos,
                'proveedores': proveedores,
                'alergenos_seleccionados': set(ing.alergenos.values_list('pk', flat=True)),
                'accion': 'Editar ingrediente',
            }))
        ing.proveedor_id     = prov_id
        ing.unidad_compra    = request.POST.get('unidad_compra', ing.unidad_compra).strip()
        ing.unidad_base      = request.POST.get('unidad_base', ing.unidad_base)

        contenido = request.POST.get('contenido_por_unidad', '').strip()
        if contenido:
            ing.contenido_por_unidad = contenido

        stock_minimo = request.POST.get('stock_minimo', '').strip()
        if stock_minimo:
            ing.stock_minimo = stock_minimo

        if request.FILES.get('imagen'):
            ing.imagen = request.FILES['imagen']
        ing.save()
        ing.alergenos.set(request.POST.getlist('alergenos'))
        messages.success(request, f'Ingrediente "{ing.nombre}" actualizado.')
        return redirect('app_admin:ingredientes')

    return render(request, 'app_admin/ingrediente_form.html', _ctx({
        'ingrediente':             ing,
        'alergenos_todos':         alergenos_todos,
        'proveedores':             proveedores,
        'alergenos_seleccionados': set(ing.alergenos.values_list('pk', flat=True)),
        'accion':                  'Editar ingrediente',
    }))

@admin_required
def ingrediente_ajuste(request, pk):
    ing = get_object_or_404(Ingrediente, pk=pk)
    if request.method == 'POST':
        tipo     = request.POST.get('tipo', 'ajuste')
        cantidad = float(request.POST.get('cantidad', 0))
        nota     = request.POST.get('nota', '').strip()

        with transaction.atomic():
            if tipo in ['salida', 'merma']:
                descontar_fifo(ing, cantidad, nota=nota)
            elif tipo == 'ajuste':
                # Manual override: set remaining stock to exact value on the oldest lot
                lote = ing.lotes.filter(
                    cantidad_base__gt=0,
                    fecha_vencimiento__gte=date.today()
                ).order_by('fecha_vencimiento').first()
                if lote:
                    lote.cantidad_base = cantidad
                    lote.save(update_fields=['cantidad_base'])

            MovimientoIngrediente.objects.create(
                ingrediente=ing, tipo=tipo, cantidad=cantidad, nota=nota
            )

        messages.success(request, f'Stock de "{ing.nombre}" actualizado. Stock real: {ing.stock_real:.2f} {ing.get_unidad_base_display()}')
    return redirect('app_admin:ingredientes')


@admin_required
def ingrediente_historial(request, pk):
    ing         = get_object_or_404(Ingrediente, pk=pk)
    movimientos = ing.movimientos.select_related('compra__proveedor').order_by('-fecha')
    return render(request, 'app_admin/ingrediente_historial.html', _ctx({
        'ingrediente': ing,
        'movimientos': movimientos,
    }))


# ══════════════════════════════════════════════════════════════════════════════
# LOTES DE INGREDIENTE
# ══════════════════════════════════════════════════════════════════════════════

@admin_required
def lotes_ingrediente(request, pk):
    ing   = get_object_or_404(Ingrediente, pk=pk)
    lotes = ing.lotes.select_related('proveedor', 'compra').order_by('fecha_vencimiento')
    form  = LoteIngredienteForm(ingrediente=ing)

    if request.method == 'POST':
        form = LoteIngredienteForm(request.POST, ingrediente=ing)
        if form.is_valid():
            with transaction.atomic():
                lote = form.save()
                MovimientoIngrediente.objects.create(
                    ingrediente=ing,
                    tipo='entrada',
                    cantidad=lote.cantidad_base,
                    nota=f'Lote manual — vence {lote.fecha_vencimiento}',
                )
            messages.success(request, f'Lote creado: {lote.cantidad_base:.2f} {ing.get_unidad_base_display()} — vence {lote.fecha_vencimiento}')
            return redirect('app_admin:lotes_ingrediente', pk=pk)
        else:
            messages.error(request, 'Corrige los errores del formulario.')

    lotes_list = list(lotes)
    lotes_activos    = sum(1 for l in lotes_list if not l.vencido and float(l.cantidad_base) > 0)
    lotes_venc_pronto = sum(1 for l in lotes_list if l.vence_pronto and float(l.cantidad_base) > 0)
    return render(request, 'app_admin/lotes_ingrediente.html', _ctx({
        'ingrediente':       ing,
        'lotes':             lotes_list,
        'form':              form,
        'proveedores':       Proveedor.objects.filter(activo=True),
        'lotes_activos':     lotes_activos,
        'lotes_venc_pronto': lotes_venc_pronto,
    }))


@admin_required
def lote_eliminar(request, pk):
    lote = get_object_or_404(LoteIngrediente, pk=pk)
    ing_pk = lote.ingrediente_id
    if request.method == 'POST':
        if float(lote.cantidad_base) > 0:
            MovimientoIngrediente.objects.create(
                ingrediente=lote.ingrediente,
                tipo='merma',
                cantidad=lote.cantidad_base,
                nota='Lote eliminado manualmente',
            )
        lote.delete()
        messages.success(request, 'Lote eliminado.')
    return redirect('app_admin:lotes_ingrediente', pk=ing_pk)


# ══════════════════════════════════════════════════════════════════════════════
# PRODUCCIÓN DE PRODUCTOS ELABORADOS
# ══════════════════════════════════════════════════════════════════════════════

@admin_required
def producciones(request):
    qs = ProduccionElaborado.objects.select_related('producto', 'responsable').order_by('-fecha')
    return render(request, 'app_admin/producciones.html', _ctx({'producciones': qs}))


@admin_required
def produccion_nueva(request):
    elaborados = Producto.objects.filter(tipo='elaborado', disponible=True).prefetch_related('receta__ingrediente')
    if request.method == 'POST':
        prod_id  = request.POST.get('producto')
        cantidad = request.POST.get('cantidad_producida', '0').strip()
        nota     = request.POST.get('nota', '').strip()

        errores = []
        if not prod_id:
            errores.append('Selecciona un producto.')
        try:
            cantidad = int(cantidad)
            if cantidad <= 0:
                errores.append('La cantidad debe ser mayor a 0.')
        except (ValueError, TypeError):
            errores.append('Cantidad inválida.')

        if not errores:
            prod = get_object_or_404(Producto, pk=prod_id, tipo='elaborado')
            nuevo_stock = None
            with transaction.atomic():
                receta_items = list(prod.receta.select_related('ingrediente').all())
                # Bug #3/#9b: producción por tandas — para ingredientes
                # con unidad_base='unidad' (huevos, etc.) la cantidad necesaria
                # debe ser un entero, si no la receta exige fraccionar lo no fraccionable.
                for receta_ing in receta_items:
                    needed = float(receta_ing.cantidad) * cantidad
                    if receta_ing.ingrediente.unidad_base == 'unidad' and abs(needed - round(needed)) > 1e-6:
                        # Calcular el menor múltiplo de tanda que daría enteros
                        cant_receta = float(receta_ing.cantidad)
                        sugerido = None
                        if cant_receta > 0:
                            for batch in range(1, 1001):
                                if abs((batch * cant_receta) - round(batch * cant_receta)) < 1e-6:
                                    sugerido = batch
                                    break
                        msg = (
                            f'No puedes producir {cantidad} unidad(es) porque exigiría '
                            f'{needed:.3f} {receta_ing.ingrediente.get_unidad_base_display()} '
                            f'de {receta_ing.ingrediente.nombre} (debe ser entero).'
                        )
                        if sugerido:
                            msg += f' Tanda mínima sugerida: múltiplos de {sugerido}.'
                        errores.append(msg)
                # Lock lotes inside transaction to prevent race conditions (Bug #8)
                # and avoid N+1 property calls (Bug #3)
                for receta_ing in receta_items:
                    needed = float(receta_ing.cantidad) * cantidad
                    lotes = LoteIngrediente.objects.select_for_update().filter(
                        ingrediente=receta_ing.ingrediente,
                        cantidad_base__gt=0,
                        fecha_vencimiento__gte=date.today()
                    )
                    disponible = sum(float(l.cantidad_base) for l in lotes)
                    if disponible < needed:
                        errores.append(
                            f'Stock insuficiente de {receta_ing.ingrediente.nombre}: '
                            f'necesitas {needed:.2f} {receta_ing.ingrediente.get_unidad_base_display()}, '
                            f'hay {disponible:.2f}.'
                        )

                if errores:
                    transaction.set_rollback(True)
                else:
                    costo_total = 0
                    for receta_ing in receta_items:
                        needed = float(receta_ing.cantidad) * cantidad
                        costo_total += needed * receta_ing.ingrediente.costo_unitario_real
                        resto = descontar_fifo(receta_ing.ingrediente, needed,
                                               nota=f'Producción {cantidad}× {prod.nombre}')
                        if resto > 0:
                            # FIFO deduction was incomplete despite validation — race condition escaped (Bug #7)
                            errores.append(
                                f'FIFO incompleto para {receta_ing.ingrediente.nombre}: '
                                f'{resto:.3f} {receta_ing.ingrediente.get_unidad_base_display()} sin cubrir.'
                            )
                            transaction.set_rollback(True)
                            break
                        MovimientoIngrediente.objects.create(
                            ingrediente=receta_ing.ingrediente,
                            tipo='salida',
                            cantidad=needed,
                            nota=f'Producción {cantidad}× {prod.nombre}',
                        )

                    if not errores:
                        prod_locked = Producto.objects.select_for_update().get(pk=prod.pk)
                        prod_locked.stock += cantidad
                        prod_locked.save(update_fields=['stock'])
                        nuevo_stock = prod_locked.stock
                        MovimientoInventario.objects.create(
                            producto=prod_locked, tipo='entrada',
                            cantidad=cantidad,
                            nota=f'Producción registrada por {request.user.get_full_name() or request.user.username}',
                        )
                        ProduccionElaborado.objects.create(
                            producto=prod_locked,
                            cantidad_producida=cantidad,
                            responsable=request.user,
                            costo_total=round(costo_total, 2),
                            nota=nota,
                        )

            if not errores:
                messages.success(request, f'Producción registrada: {cantidad}× {prod.nombre}. Stock ahora: {nuevo_stock}.')
                return redirect('app_admin:producciones')

        for e in errores:
            messages.error(request, e)

    # Build ingredient summary for each elaborated product (for JS preview)
    prods_data = []
    for p in elaborados:
        receta = [
            {
                'nombre': r.ingrediente.nombre,
                'cantidad': float(r.cantidad),
                'unidad': r.ingrediente.get_unidad_base_display(),
                'stock_real': r.ingrediente.stock_real,
                # Bug #3/#9b: marcamos los ingredientes que solo aceptan unidades enteras
                'unitario': r.ingrediente.unidad_base == 'unidad',
            }
            for r in p.receta.select_related('ingrediente').all()
        ]
        max_producible = min(
            (r['stock_real'] // r['cantidad'] if r['cantidad'] else 0)
            for r in receta
        ) if receta else 0
        # Bug #3/#9b: tanda mínima — menor batch entero tal que cant*recetas dé enteros
        # en TODOS los ingredientes "unitarios" simultáneamente
        unitarios = [r for r in receta if r['unitario'] and r['cantidad'] > 0]
        tanda_min = 1
        if unitarios:
            for b in range(1, 1001):
                if all(abs((b * r['cantidad']) - round(b * r['cantidad'])) < 1e-6 for r in unitarios):
                    tanda_min = b
                    break
        prods_data.append({
            'id': p.pk,
            'nombre': p.nombre,
            'stock': p.stock,
            'receta': receta,
            'max_producible': int(max_producible),
            'tanda_min': tanda_min,
        })

    return render(request, 'app_admin/produccion_form.html', _ctx({
        'elaborados':  elaborados,
        'prods_json':  json.dumps(prods_data),
    }))


@admin_required
def exportar_ingredientes_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messages.error(request, 'openpyxl no está instalado.')
        return redirect('app_admin:ingredientes')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ingredientes'

    header_fill = PatternFill(start_color='1a3d2b', end_color='1a3d2b', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    headers = ['Nombre', 'Proveedor', 'Unidad compra', 'Contenido/ud', 'Unidad base', 'Stock real', 'Mínimo (base)', 'Costo unitario', 'Próx. vencimiento', 'Lotes activos']
    ws.append(headers)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for ing in Ingrediente.objects.filter(activo=True).select_related('proveedor').prefetch_related('lotes').order_by('nombre'):
        lote_prox = ing.lote_proximo_vencer
        ws.append([
            ing.nombre,
            ing.proveedor.nombre if ing.proveedor else '—',
            ing.unidad_compra,
            float(ing.contenido_por_unidad),
            ing.get_unidad_base_display(),
            ing.stock_real,
            float(ing.stock_minimo),
            ing.costo_unitario_real,
            lote_prox.fecha_vencimiento.strftime('%d/%m/%Y') if lote_prox else '—',
            ing.lotes.filter(cantidad_base__gt=0).count(),
        ])

    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="ingredientes_punto_asis.xlsx"'
    wb.save(response)
    return response


@admin_required
def exportar_ingredientes_pdf(request):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        import io
    except ImportError:
        messages.error(request, 'reportlab no está instalado.')
        return redirect('app_admin:ingredientes')

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph('Ingredientes — Punto Asis', styles['Title']))
    elements.append(Paragraph(f'Generado: {date.today().strftime("%d/%m/%Y")}', styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))

    headers = ['Nombre', 'Proveedor', 'Ud. compra', 'Cont./ud', 'Ud. base', 'Stock real', 'Mínimo', 'Costo/ud base', 'Próx. vencimiento', 'Lotes']
    data = [headers]
    for ing in Ingrediente.objects.filter(activo=True).select_related('proveedor').prefetch_related('lotes').order_by('nombre'):
        lote_prox = ing.lote_proximo_vencer
        data.append([
            ing.nombre[:28],
            (ing.proveedor.nombre if ing.proveedor else '—')[:18],
            ing.unidad_compra,
            f'{float(ing.contenido_por_unidad):g}',
            ing.get_unidad_base_display(),
            f'{ing.stock_real:g}',
            f'{float(ing.stock_minimo):g}',
            f'${ing.costo_unitario_real:,.4f}',
            lote_prox.fecha_vencimiento.strftime('%d/%m/%Y') if lote_prox else '—',
            str(ing.lotes.filter(cantidad_base__gt=0).count()),
        ])

    verde_oscuro = colors.HexColor('#1a3d2b')
    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), verde_oscuro),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0fdf4')]),
        ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#e2e2dd')),
        ('ALIGN',      (2, 0), (4, -1), 'RIGHT'),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(tabla)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="ingredientes_punto_asis.pdf"'
    return response


# ══════════════════════════════════════════════════════════════════════════════
# HISTORIAL UNIFICADO
# ══════════════════════════════════════════════════════════════════════════════

@admin_required
def historial(request):
    return render(request, 'app_admin/historial.html', _ctx())


def historial_api(request):
    """API JSON para el historial unificado con filtros.
       Convierte las fechas UTC a hora local de Colombia (America/Bogota).
    """
    tipo_f  = request.GET.get('tipo', '')     # 'producto' | 'ingrediente'
    mov_f   = request.GET.get('mov', '')      # entrada | salida | ajuste | merma
    desde   = request.GET.get('desde', '')
    hasta   = request.GET.get('hasta', '')
    q       = request.GET.get('q', '').strip()

    resultados = []

    # Movimientos de productos
    if tipo_f in ('', 'producto'):
        qs = MovimientoInventario.objects.select_related(
            'producto__proveedor', 'compra__proveedor'
        ).order_by('-fecha')
        if mov_f:   qs = qs.filter(tipo=mov_f)
        if desde:   qs = qs.filter(fecha__date__gte=desde)
        if hasta:   qs = qs.filter(fecha__date__lte=hasta)
        if q:       qs = qs.filter(producto__nombre__icontains=q)
        for m in qs[:200]:
            # Proveedor: 1) el de la compra vinculada, 2) el registrado en el producto
            if m.compra and m.compra.proveedor:
                prov = m.compra.proveedor.nombre
            elif m.producto.proveedor:
                prov = m.producto.proveedor.nombre
            else:
                prov = '—'
            # Convertir UTC a hora local
            local_fecha = timezone.localtime(m.fecha)
            resultados.append({
                'categoria': 'Producto',
                'nombre':    m.producto.nombre,
                'tipo':      m.get_tipo_display(),
                'tipo_raw':  m.tipo,
                'cantidad':  float(m.cantidad),
                'proveedor': prov,
                'nota':      m.nota,
                'fecha':     local_fecha.strftime('%d/%m/%Y %H:%M'),
                'fecha_iso': local_fecha.isoformat(),
                'fecha_timestamp': int(local_fecha.timestamp() * 1000),  # ← Nuevo campo
            })

    # Movimientos de ingredientes
    if tipo_f in ('', 'ingrediente'):
        qs = MovimientoIngrediente.objects.select_related(
            'ingrediente__proveedor', 'compra__proveedor'
        ).order_by('-fecha')
        if mov_f:   qs = qs.filter(tipo=mov_f)
        if desde:   qs = qs.filter(fecha__date__gte=desde)
        if hasta:   qs = qs.filter(fecha__date__lte=hasta)
        if q:       qs = qs.filter(ingrediente__nombre__icontains=q)
        for m in qs[:200]:
            if m.compra and m.compra.proveedor:
                prov = m.compra.proveedor.nombre
            elif hasattr(m.ingrediente, 'proveedor') and m.ingrediente.proveedor:
                prov = m.ingrediente.proveedor.nombre
            else:
                prov = '—'
            # Convertir UTC a hora local
            local_fecha = timezone.localtime(m.fecha)
            resultados.append({
                'categoria': 'Ingrediente',
                'nombre':    m.ingrediente.nombre,
                'tipo':      m.get_tipo_display(),
                'tipo_raw':  m.tipo,
                'cantidad':  float(m.cantidad),
                'proveedor': prov,
                'nota':      m.nota,
                'fecha':     local_fecha.strftime('%d/%m/%Y %H:%M'),
                'fecha_iso': local_fecha.isoformat(),
                'fecha_timestamp': int(local_fecha.timestamp() * 1000),  # ← Nuevo campo
            })

    # Ordenar por fecha ISO (ahora local, pero sigue siendo ordenable)
    resultados.sort(key=lambda x: x['fecha_iso'], reverse=True)
    return JsonResponse({'data': resultados[:300]})

# ══════════════════════════════════════════════════════════════════════════════
# CALENDARIO DE MOVIMIENTOS
# ══════════════════════════════════════════════════════════════════════════════

@admin_required
def calendario(request):
    token    = _ical_token()
    ical_url = request.build_absolute_uri(
        reverse('app_admin:calendario_ical', args=[token])
    )
    padmin, _ = PerfilAdmin.objects.get_or_create(perfil=request.user.perfil)
    try:
        gcal_token = padmin.gcal
    except GoogleCalendarToken.DoesNotExist:
        gcal_token = None
    return render(request, 'app_admin/calendario.html', _ctx({
        'ical_url':   ical_url,
        'gcal_token': gcal_token,
    }))


@admin_required
def calendario_api(request):
    """Devuelve movimientos agrupados por fecha para FullCalendar.
       Convierte fechas UTC a hora local de Colombia (America/Bogota).
    """
    desde = request.GET.get('start', str(date.today().replace(day=1)))[:10]
    hasta = request.GET.get('end',   str(date.today()))[:10]

    eventos = []

    # Entradas de productos
    for m in MovimientoInventario.objects.filter(
        fecha__date__gte=desde, fecha__date__lte=hasta, tipo='entrada'
    ).select_related('producto'):
        fecha_local = timezone.localtime(m.fecha).date()
        eventos.append({
            'title':           f'+{int(m.cantidad)} {m.producto.nombre}',
            'start':           fecha_local.isoformat(),
            'backgroundColor': '#2d6a4f',
            'borderColor':     '#2d6a4f',
            'textColor':       '#fff',
            'extendedProps':   {'tipo': 'entrada', 'cat': 'Producto'},
        })

    # Salidas de productos
    for m in MovimientoInventario.objects.filter(
        fecha__date__gte=desde, fecha__date__lte=hasta, tipo__in=['salida', 'merma']
    ).select_related('producto'):
        fecha_local = timezone.localtime(m.fecha).date()
        eventos.append({
            'title':           f'-{int(m.cantidad)} {m.producto.nombre}',
            'start':           fecha_local.isoformat(),
            'backgroundColor': '#d94f3d',
            'borderColor':     '#d94f3d',
            'textColor':       '#fff',
            'extendedProps':   {'tipo': m.tipo, 'cat': 'Producto'},
        })

    # Movimientos de ingredientes
    for m in MovimientoIngrediente.objects.filter(
        fecha__date__gte=desde, fecha__date__lte=hasta
    ).select_related('ingrediente'):
        color = '#1a56db' if m.tipo == 'entrada' else '#e07b2a'
        fecha_local = timezone.localtime(m.fecha).date()
        eventos.append({
            'title':           f'Ing: {m.ingrediente.nombre}',
            'start':           fecha_local.isoformat(),
            'backgroundColor': color,
            'borderColor':     color,
            'textColor':       '#fff',
            'extendedProps':   {'tipo': m.tipo, 'cat': 'Ingrediente'},
        })

    # Pedidos estudiante entregados
    for p in Pedido.objects.filter(
        fecha_pedido__date__gte=desde, fecha_pedido__date__lte=hasta, estado='entregado'
    ).select_related('estudiante__perfil__user'):
        fecha_local = timezone.localtime(p.fecha_pedido).date()
        nombre = p.estudiante.perfil.user.get_full_name() or p.estudiante.perfil.user.username
        eventos.append({
            'title':           f'{p.ticket} — {nombre}',
            'start':           fecha_local.isoformat(),
            'backgroundColor': '#7c3aed',
            'borderColor':     '#7c3aed',
            'textColor':       '#fff',
            'extendedProps':   {'tipo': 'entregado', 'cat': 'Estudiante'},
        })

    # Pedidos docente entregados
    for p in PedidoDocente.objects.filter(
        fecha_pedido__date__gte=desde, fecha_pedido__date__lte=hasta, estado='entregado'
    ).select_related('docente__perfil__user'):
        fecha_local = timezone.localtime(p.fecha_pedido).date()
        nombre = p.docente.perfil.user.get_full_name() or p.docente.perfil.user.username
        eventos.append({
            'title':           f'{p.ticket} — {nombre}',
            'start':           fecha_local.isoformat(),
            'backgroundColor': '#ea580c',
            'borderColor':     '#ea580c',
            'textColor':       '#fff',
            'extendedProps':   {'tipo': 'entregado', 'cat': 'Docente'},
        })

    return JsonResponse(eventos, safe=False)


def calendario_ical(request, token):
    """Feed iCal de movimientos — accesible sin sesión mediante token."""
    if token != _ical_token():
        return HttpResponse('Acceso no autorizado.', status=401,
                            content_type='text/plain; charset=utf-8')

    hoy   = date.today()
    desde = (hoy - timedelta(days=60)).isoformat()
    hasta = (hoy + timedelta(days=60)).isoformat()

    def esc(s):
        return str(s).replace('\\', '\\\\').replace(';', '\\;').replace(',', '\\,').replace('\n', '\\n')

    def vdate(d):
        return d.strftime('%Y%m%d')

    lines = [
        'BEGIN:VCALENDAR',
        'VERSION:2.0',
        'PRODID:-//Punto Asis//Cafeteria Escolar//ES',
        'CALSCALE:GREGORIAN',
        'METHOD:PUBLISH',
        'X-WR-CALNAME:Punto Asis',
        'X-WR-TIMEZONE:America/Bogota',
    ]

    # Entradas de producto
    for m in MovimientoInventario.objects.filter(
        fecha__date__gte=desde, fecha__date__lte=hasta, tipo='entrada'
    ).select_related('producto').order_by('fecha'):
        d = timezone.localtime(m.fecha).date()
        lines += [
            'BEGIN:VEVENT',
            f'UID:prod-ent-{m.pk}@puntoasis',
            f'DTSTART;VALUE=DATE:{vdate(d)}',
            f'DTEND;VALUE=DATE:{vdate(d)}',
            f'SUMMARY:{esc(f"Entrada: {m.producto.nombre} +{int(m.cantidad)}")}',
            'CATEGORIES:Inventario',
            'END:VEVENT',
        ]

    # Salidas / merma de producto
    for m in MovimientoInventario.objects.filter(
        fecha__date__gte=desde, fecha__date__lte=hasta, tipo__in=['salida', 'merma']
    ).select_related('producto').order_by('fecha'):
        d = timezone.localtime(m.fecha).date()
        lines += [
            'BEGIN:VEVENT',
            f'UID:prod-sal-{m.pk}@puntoasis',
            f'DTSTART;VALUE=DATE:{vdate(d)}',
            f'DTEND;VALUE=DATE:{vdate(d)}',
            f'SUMMARY:{esc(f"{m.get_tipo_display()}: {m.producto.nombre} -{int(m.cantidad)}")}',
            'CATEGORIES:Inventario',
            'END:VEVENT',
        ]

    # Movimientos de ingredientes
    for m in MovimientoIngrediente.objects.filter(
        fecha__date__gte=desde, fecha__date__lte=hasta
    ).select_related('ingrediente').order_by('fecha'):
        d = timezone.localtime(m.fecha).date()
        lines += [
            'BEGIN:VEVENT',
            f'UID:ing-{m.tipo}-{m.pk}@puntoasis',
            f'DTSTART;VALUE=DATE:{vdate(d)}',
            f'DTEND;VALUE=DATE:{vdate(d)}',
            f'SUMMARY:{esc(f"Ing {m.get_tipo_display()}: {m.ingrediente.nombre}")}',
            'CATEGORIES:Ingredientes',
            'END:VEVENT',
        ]

    # Pedidos estudiante entregados
    for p in Pedido.objects.filter(
        fecha_pedido__date__gte=desde, fecha_pedido__date__lte=hasta, estado='entregado'
    ).select_related('estudiante__perfil__user').order_by('fecha_pedido'):
        d = timezone.localtime(p.fecha_pedido).date()
        nombre = p.estudiante.perfil.user.get_full_name() if p.estudiante_id else 'Estudiante'
        lines += [
            'BEGIN:VEVENT',
            f'UID:pedido-{p.pk}@puntoasis',
            f'DTSTART;VALUE=DATE:{vdate(d)}',
            f'DTEND;VALUE=DATE:{vdate(d)}',
            f'SUMMARY:{esc(f"Pedido {p.ticket} — {nombre}")}',
            'CATEGORIES:Pedidos',
            'END:VEVENT',
        ]

    # Pedidos docente entregados
    for p in PedidoDocente.objects.filter(
        fecha_pedido__date__gte=desde, fecha_pedido__date__lte=hasta, estado='entregado'
    ).select_related('docente__perfil__user').order_by('fecha_pedido'):
        d = timezone.localtime(p.fecha_pedido).date()
        nombre = p.docente.perfil.user.get_full_name() if p.docente_id else 'Docente'
        lines += [
            'BEGIN:VEVENT',
            f'UID:pedido-doc-{p.pk}@puntoasis',
            f'DTSTART;VALUE=DATE:{vdate(d)}',
            f'DTEND;VALUE=DATE:{vdate(d)}',
            f'SUMMARY:{esc(f"Pedido docente {p.ticket} — {nombre}")}',
            'CATEGORIES:Pedidos',
            'END:VEVENT',
        ]

    lines.append('END:VCALENDAR')
    content = '\r\n'.join(lines) + '\r\n'

    response = HttpResponse(content, content_type='text/calendar; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="puntoasis.ics"'
    return response


# ══════════════════════════════════════════════════════════════════════════════
# GOOGLE CALENDAR — OAuth2 + sync
# ══════════════════════════════════════════════════════════════════════════════

_GCAL_SCOPES = ['https://www.googleapis.com/auth/calendar']

def _gcal_flow(redirect_uri, state=None):
    """Construye un Flow de google-auth-oauthlib."""
    from google_auth_oauthlib.flow import Flow
    cfg = {
        'web': {
            'client_id':     settings.GOOGLE_CLIENT_ID,
            'client_secret': settings.GOOGLE_CLIENT_SECRET,
            'auth_uri':      'https://accounts.google.com/o/oauth2/auth',
            'token_uri':     'https://oauth2.googleapis.com/token',
            'redirect_uris': [redirect_uri],
        }
    }
    kwargs = {'state': state} if state else {}
    flow = Flow.from_client_config(cfg, scopes=_GCAL_SCOPES, **kwargs)
    flow.redirect_uri = redirect_uri
    return flow


@admin_required
def calendario_google_auth(request):
    """Inicia el flujo OAuth2 para conectar Google Calendar."""
    import secrets
    try:
        from google_auth_oauthlib.flow import Flow  # noqa — solo para validar instalación
    except ImportError:
        messages.error(request, 'Paquetes de Google no instalados. Ejecuta: pip install google-auth google-auth-oauthlib google-api-python-client')
        return redirect('app_admin:calendario')

    site_url = settings.SITE_URL.rstrip('/')
    redirect_uri = site_url + reverse('app_admin:calendario_google_callback')
    state = secrets.token_urlsafe(16)
    request.session['gcal_state']        = state
    request.session['gcal_redirect_uri'] = redirect_uri

    import hashlib, base64
    code_verifier  = secrets.token_urlsafe(64)
    code_challenge = base64.urlsafe_b64encode(
        hashlib.sha256(code_verifier.encode()).digest()
    ).rstrip(b'=').decode()
    request.session['gcal_code_verifier'] = code_verifier

    flow = _gcal_flow(redirect_uri)
    auth_url, _ = flow.authorization_url(
        access_type='offline',
        prompt='consent',
        state=state,
        include_granted_scopes='false',
        code_challenge=code_challenge,
        code_challenge_method='S256',
    )
    return redirect(auth_url)


@admin_required
def calendario_google_callback(request):
    """Recibe el callback de Google y guarda el token."""
    state        = request.session.get('gcal_state')
    redirect_uri = request.session.get('gcal_redirect_uri')

    if not state or request.GET.get('state') != state or not redirect_uri:
        messages.error(request, 'Estado OAuth inválido. Intenta de nuevo.')
        return redirect('app_admin:calendario')

    if 'error' in request.GET:
        messages.error(request, f'Google rechazó el acceso: {request.GET["error"]}')
        return redirect('app_admin:calendario')

    code_verifier = request.session.get('gcal_code_verifier')
    try:
        flow = _gcal_flow(redirect_uri, state=state)
        flow.fetch_token(code=request.GET.get('code'), code_verifier=code_verifier)
    except Exception as e:
        import traceback; traceback.print_exc()
        messages.error(request, f'Error al obtener el token de Google: {e}')
        return redirect('app_admin:calendario')

    creds  = flow.credentials
    padmin, _ = PerfilAdmin.objects.get_or_create(perfil=request.user.perfil)
    token, _  = GoogleCalendarToken.objects.get_or_create(admin=padmin)
    token.access_token  = creds.token
    if creds.refresh_token:
        token.refresh_token = creds.refresh_token
    token.token_expiry  = creds.expiry
    token.gcal_id       = 'primary'   # se asigna en el primer sync
    token.save()

    request.session.pop('gcal_state', None)
    request.session.pop('gcal_redirect_uri', None)
    request.session.pop('gcal_code_verifier', None)

    messages.success(request, 'Google Calendar conectado. Haz clic en Sincronizar para enviar los eventos.')
    return redirect('app_admin:calendario')


@admin_required
def calendario_google_sync(request):
    """Sincroniza eventos del app con Google Calendar (POST)."""
    if request.method != 'POST':
        return redirect('app_admin:calendario')

    try:
        from google.oauth2.credentials import Credentials
        from google.auth.transport.requests import Request as GRequest
        from googleapiclient.discovery import build
        from googleapiclient.errors import HttpError
    except ImportError:
        messages.error(request, 'Paquetes de Google no instalados. Ejecuta: pip install google-auth google-auth-oauthlib google-api-python-client')
        return redirect('app_admin:calendario')

    padmin, _ = PerfilAdmin.objects.get_or_create(perfil=request.user.perfil)
    try:
        gtoken = padmin.gcal
    except GoogleCalendarToken.DoesNotExist:
        messages.error(request, 'Conecta Google Calendar primero.')
        return redirect('app_admin:calendario')

    # Construir credenciales y refrescar si es necesario
    creds = Credentials(
        token=gtoken.access_token,
        refresh_token=gtoken.refresh_token or None,
        token_uri='https://oauth2.googleapis.com/token',
        client_id=settings.GOOGLE_CLIENT_ID,
        client_secret=settings.GOOGLE_CLIENT_SECRET,
        scopes=_GCAL_SCOPES,
    )
    if creds.expired and creds.refresh_token:
        try:
            creds.refresh(GRequest())
            gtoken.access_token = creds.token
            gtoken.token_expiry = creds.expiry
            gtoken.save(update_fields=['access_token', 'token_expiry'])
        except Exception:
            messages.error(request, 'El token de Google expiró. Reconecta tu cuenta.')
            gtoken.delete()
            return redirect('app_admin:calendario')

    service = build('calendar', 'v3', credentials=creds)

    # Crear o recuperar el calendario "Punto Asis"
    if not gtoken.gcal_id or gtoken.gcal_id == 'primary':
        try:
            cal_list  = service.calendarList().list().execute()
            asis_cal  = next(
                (c for c in cal_list.get('items', []) if c.get('summary') == 'Punto Asis'),
                None,
            )
            if asis_cal:
                gtoken.gcal_id = asis_cal['id']
            else:
                new_cal = service.calendars().insert(body={
                    'summary':     'Punto Asis',
                    'description': 'Eventos de la cafetería escolar Punto Asis',
                    'timeZone':    'America/Bogota',
                }).execute()
                gtoken.gcal_id = new_cal['id']
            gtoken.save(update_fields=['gcal_id'])
        except Exception:
            messages.error(request, 'No se pudo crear el calendario en Google. Verifica los permisos.')
            return redirect('app_admin:calendario')

    gcal_id = gtoken.gcal_id
    hoy   = date.today()
    desde = (hoy - timedelta(days=30)).isoformat()
    hasta = (hoy + timedelta(days=60)).isoformat()

    # Leer eventos existentes de Google Calendar marcados con source=puntoasis
    existing = {}   # uid → google event id
    page_token = None
    while True:
        try:
            result = service.events().list(
                calendarId=gcal_id,
                timeMin=f'{desde}T00:00:00-05:00',
                timeMax=f'{hasta}T23:59:59-05:00',
                privateExtendedProperty='source=puntoasis',
                pageToken=page_token,
                maxResults=500,
            ).execute()
        except Exception:
            break
        for ev in result.get('items', []):
            uid = ev.get('extendedProperties', {}).get('private', {}).get('uid', '')
            if uid:
                existing[uid] = ev['id']
        page_token = result.get('nextPageToken')
        if not page_token:
            break

    # Construir lista de eventos del app
    COLOR = {
        'prod_ent': '2',   # sage
        'prod_sal': '11',  # tomato
        'ing_ent':  '7',   # peacock
        'ing_sal':  '6',   # tangerine
        'pedido':   '3',   # grape
        'ped_doc':  '4',   # flamingo
    }

    app_events = []
    for m in MovimientoInventario.objects.filter(
        fecha__date__gte=desde, fecha__date__lte=hasta, tipo='entrada'
    ).select_related('producto'):
        d = timezone.localtime(m.fecha).date()
        app_events.append({'uid': f'prod-ent-{m.pk}', 'colorId': COLOR['prod_ent'],
                           'summary': f'Entrada: {m.producto.nombre} +{int(m.cantidad)}',
                           'date': d.isoformat()})

    for m in MovimientoInventario.objects.filter(
        fecha__date__gte=desde, fecha__date__lte=hasta, tipo__in=['salida', 'merma']
    ).select_related('producto'):
        d = timezone.localtime(m.fecha).date()
        app_events.append({'uid': f'prod-sal-{m.pk}', 'colorId': COLOR['prod_sal'],
                           'summary': f'{m.get_tipo_display()}: {m.producto.nombre} -{int(m.cantidad)}',
                           'date': d.isoformat()})

    for m in MovimientoIngrediente.objects.filter(
        fecha__date__gte=desde, fecha__date__lte=hasta
    ).select_related('ingrediente'):
        d    = timezone.localtime(m.fecha).date()
        key  = 'ing_ent' if m.tipo == 'entrada' else 'ing_sal'
        app_events.append({'uid': f'ing-{m.tipo}-{m.pk}', 'colorId': COLOR[key],
                           'summary': f'Ing {m.get_tipo_display()}: {m.ingrediente.nombre}',
                           'date': d.isoformat()})

    for p in Pedido.objects.filter(
        fecha_pedido__date__gte=desde, fecha_pedido__date__lte=hasta, estado='entregado'
    ).select_related('estudiante__perfil__user'):
        d      = timezone.localtime(p.fecha_pedido).date()
        nombre = p.estudiante.perfil.user.get_full_name() if p.estudiante_id else 'Estudiante'
        app_events.append({'uid': f'pedido-{p.pk}', 'colorId': COLOR['pedido'],
                           'summary': f'Pedido {p.ticket} — {nombre}',
                           'date': d.isoformat()})

    for p in PedidoDocente.objects.filter(
        fecha_pedido__date__gte=desde, fecha_pedido__date__lte=hasta, estado='entregado'
    ).select_related('docente__perfil__user'):
        d      = timezone.localtime(p.fecha_pedido).date()
        nombre = p.docente.perfil.user.get_full_name() if p.docente_id else 'Docente'
        app_events.append({'uid': f'ped-doc-{p.pk}', 'colorId': COLOR['ped_doc'],
                           'summary': f'Pedido docente {p.ticket} — {nombre}',
                           'date': d.isoformat()})

    app_uids = {e['uid'] for e in app_events}
    created  = 0
    deleted  = 0

    for ev in app_events:
        if ev['uid'] not in existing:
            try:
                service.events().insert(
                    calendarId=gcal_id,
                    body={
                        'summary': ev['summary'],
                        'start':   {'date': ev['date']},
                        'end':     {'date': ev['date']},
                        'colorId': ev['colorId'],
                        'extendedProperties': {
                            'private': {'source': 'puntoasis', 'uid': ev['uid']}
                        },
                    },
                ).execute()
                created += 1
            except HttpError:
                pass

    for uid, gev_id in existing.items():
        if uid not in app_uids:
            try:
                service.events().delete(calendarId=gcal_id, eventId=gev_id).execute()
                deleted += 1
            except HttpError:
                pass

    gtoken.synced_at = timezone.now()
    gtoken.save(update_fields=['synced_at'])

    messages.success(request, f'Sincronizado: {created} evento(s) creado(s), {deleted} eliminado(s).')
    return redirect('app_admin:calendario')


@admin_required
def calendario_google_disconnect(request):
    """Elimina la conexión con Google Calendar (POST)."""
    if request.method != 'POST':
        return redirect('app_admin:calendario')
    padmin, _ = PerfilAdmin.objects.get_or_create(perfil=request.user.perfil)
    try:
        padmin.gcal.delete()
        messages.success(request, 'Google Calendar desconectado.')
    except GoogleCalendarToken.DoesNotExist:
        pass
    return redirect('app_admin:calendario')


# ══════════════════════════════════════════════════════════════════════════════
# PROVEEDORES
# ══════════════════════════════════════════════════════════════════════════════

@admin_required
def proveedores(request):
    q  = request.GET.get('q', '').strip()
    qs = Proveedor.objects.annotate(
        num_productos=Count('productos', distinct=True),
        num_insumos=Count('insumos', distinct=True),
        num_compras=Count('compras', distinct=True),
    ).order_by('nombre')
    if q:
        qs = qs.filter(Q(nombre__icontains=q) | Q(nit__icontains=q) | Q(contacto__icontains=q))

    # Build items JSON for each proveedor (for modal)
    proveedores_data = []
    for prov in qs:
        items = []
        for p in Producto.objects.filter(proveedor=prov, tipo='simple'):
            items.append({'tipo': 'Producto', 'nombre': p.nombre, 'stock': str(p.stock), 'unidad': 'und'})
        for ing in Ingrediente.objects.filter(proveedor=prov, activo=True):
            items.append({'tipo': 'Ingrediente', 'nombre': ing.nombre, 'stock': f'{ing.stock_real:.2f}', 'unidad': ing.get_unidad_base_display()})
        for ins in Insumo.objects.filter(proveedor=prov, activo=True):
            items.append({'tipo': 'Insumo', 'nombre': ins.nombre, 'stock': str(ins.stock), 'unidad': ins.get_unidad_display()})
        proveedores_data.append({'prov': prov, 'items_json': json.dumps(items)})

    return render(request, 'app_admin/proveedores.html', _ctx({
        'proveedores': qs,
        'proveedores_data': proveedores_data,
        'q': q,
    }))


@admin_required
def proveedor_nuevo(request):
    if request.method == 'POST':
        nombre    = request.POST.get('nombre', '').strip()
        if not nombre:
            messages.error(request, 'El nombre es obligatorio.')
        else:
            p = Proveedor.objects.create(
                nombre=nombre,
                nit=request.POST.get('nit', '').strip(),
                contacto=request.POST.get('contacto', '').strip(),
                telefono=request.POST.get('telefono', '').strip(),
                email=request.POST.get('email', '').strip(),
                direccion=request.POST.get('direccion', '').strip(),
                logo=request.FILES.get('logo'),
            )
            messages.success(request, f'Proveedor "{p.nombre}" creado.')
            return redirect('app_admin:proveedores')
    return render(request, 'app_admin/proveedor_form.html', _ctx({'accion': 'Nuevo proveedor'}))


@admin_required
def proveedor_editar(request, pk):
    prov = get_object_or_404(Proveedor, pk=pk)
    if request.method == 'POST':
        prov.nombre    = request.POST.get('nombre', prov.nombre).strip()
        prov.nit       = request.POST.get('nit', '').strip()
        prov.contacto  = request.POST.get('contacto', '').strip()
        prov.telefono  = request.POST.get('telefono', '').strip()
        prov.email     = request.POST.get('email', '').strip()
        prov.direccion = request.POST.get('direccion', '').strip()
        if request.FILES.get('logo'):
            prov.logo = request.FILES['logo']
        prov.save()
        messages.success(request, f'Proveedor "{prov.nombre}" actualizado.')
        return redirect('app_admin:proveedores')
    return render(request, 'app_admin/proveedor_form.html', _ctx({'proveedor': prov, 'accion': 'Editar proveedor'}))


@admin_required
def proveedor_compras(request, pk):
    prov    = get_object_or_404(Proveedor, pk=pk)
    compras = prov.compras.prefetch_related(
        'detalles__producto', 'detalles__ingrediente'
    ).order_by('-fecha')
    return render(request, 'app_admin/proveedor_compras.html', _ctx({
        'proveedor': prov, 'compras': compras,
    }))


@admin_required
def proveedor_eliminar(request, pk):
    prov = get_object_or_404(Proveedor, pk=pk)
    if request.method == 'POST':
        nombre = prov.nombre
        prov.activo = False
        prov.save(update_fields=['activo'])
        messages.success(request, f'Proveedor "{nombre}" desactivado.')
    return redirect('app_admin:proveedores')


@admin_required
def proveedor_stats(request, pk):
    """Estadísticas mensuales del proveedor."""
    prov = get_object_or_404(Proveedor, pk=pk)
    # Compras de los últimos 12 meses agrupadas por mes
    from django.db.models.functions import TruncMonth
    datos = prov.compras.annotate(mes=TruncMonth('fecha')).values('mes').annotate(
        total=Sum('total')
    ).order_by('mes')
    return JsonResponse({
        'labels': [d['mes'].strftime('%b %Y') for d in datos],
        'values': [float(d['total']) for d in datos],
    })


@admin_required
def compra_nueva(request, prov_pk=None):
    proveedores  = Proveedor.objects.filter(activo=True)
    productos    = Producto.objects.filter(tipo='simple')
    ingredientes = Ingrediente.objects.filter(activo=True)
    prov_sel     = get_object_or_404(Proveedor, pk=prov_pk) if prov_pk else None

    if request.method == 'POST':
        # Redirigir a la entrada unificada con los mismos datos
        return redirect('app_admin:entrada_stock')

    return render(request, 'app_admin/entrada_stock.html', _ctx({
        'proveedores': proveedores, 'productos': productos,
        'ingredientes': ingredientes, 'prov_sel': prov_sel,
    }))


# ══════════════════════════════════════════════════════════════════════════════
# PEDIDOS
# ══════════════════════════════════════════════════════════════════════════════

@admin_required
def pedidos(request):
    hoy   = date.today()
    fecha = request.GET.get('fecha', str(hoy))
    q     = request.GET.get('q', '').strip()

    qs_est = list(
        Pedido.objects.filter(fecha_pedido__date=fecha)
        .select_related('estudiante__perfil__user')
        .prefetch_related('detalles__producto')
    )
    qs_doc = list(
        PedidoDocente.objects.filter(fecha_pedido__date=fecha)
        .select_related('docente__perfil__user')
        .prefetch_related('detalles__producto')
    )

    for p in qs_est:
        p.tipo       = 'estudiante'
        p.detalle_url = reverse('app_admin:pedido_detalle', args=[p.pk])
        p.estado_url  = reverse('app_admin:pedido_estado', args=[p.pk])
        p.nombre      = p.estudiante.perfil.user.get_full_name() or p.estudiante.perfil.user.email
        p.subtitulo   = f'Grado {p.estudiante.grado}'

    for p in qs_doc:
        p.tipo        = 'docente'
        p.detalle_url = reverse('app_admin:pedido_docente_detalle', args=[p.pk])
        p.estado_url  = reverse('app_admin:pedido_docente_estado', args=[p.pk])
        p.nombre      = p.docente.perfil.user.get_full_name() or p.docente.perfil.user.email
        p.subtitulo   = p.docente.materia

    todos = sorted(qs_est + qs_doc, key=lambda p: p.fecha_pedido)

    if q:
        ql = q.lower()
        todos = [p for p in todos if ql in p.ticket.lower() or ql in p.nombre.lower()]

    kanban = {
        'pendiente':  [p for p in todos if p.estado == 'pendiente'],
        'preparando': [p for p in todos if p.estado == 'preparando'],
        'listo':      [p for p in todos if p.estado == 'listo'],
        'entregado':  [p for p in todos if p.estado == 'entregado'],
        'cancelado':  [p for p in todos if p.estado == 'cancelado'],
    }
    return render(request, 'app_admin/pedidos.html', _ctx({
        'kanban':    kanban,
        'pedidos':   todos,
        'fecha':     fecha,
        'fecha_hoy': hoy,
        'q':         q,
    }))


@admin_required
def pedido_detalle(request, pk):
    pedido = get_object_or_404(
        Pedido.objects.select_related('estudiante__perfil__user')
                      .prefetch_related('detalles__producto'),
        pk=pk
    )
    return render(request, 'app_admin/pedido_detalle.html', _ctx({'pedido': pedido}))


TRANSICIONES_VALIDAS = {
    'pendiente':  ['preparando', 'cancelado'],
    'preparando': ['listo', 'pendiente', 'cancelado'],
    'listo':      ['entregado', 'pendiente'],
    'entregado':  ['cancelado'],  # Bug #1: permite anular una entrega y restaurar stock
    'cancelado':  [],
}


@admin_required
def pedido_estado(request, pk):
    pedido = get_object_or_404(Pedido, pk=pk)
    next_url = request.POST.get('next', 'app_admin:pedidos')
    if request.method == 'POST':
        nuevo = request.POST.get('estado')
        estados_validos = [e[0] for e in Pedido.ESTADO_CHOICES]
        if nuevo not in estados_validos:
            messages.error(request, 'Estado no válido.')
            return redirect(next_url)

        if nuevo not in TRANSICIONES_VALIDAS.get(pedido.estado, []):
            messages.error(
                request,
                f'No se puede cambiar de "{pedido.get_estado_display()}" a "{dict(Pedido.ESTADO_CHOICES).get(nuevo)}".'
            )
            return redirect(next_url)

        with transaction.atomic():
            pedido_locked = Pedido.objects.select_for_update().get(pk=pk)

            # Re-validar transición con datos frescos
            if nuevo not in TRANSICIONES_VALIDAS.get(pedido_locked.estado, []):
                messages.error(request, 'Transición de estado inválida.')
                return redirect(next_url)

            pedido_locked.estado = nuevo
            if nuevo == 'entregado':
                pedido_locked.fecha_entrega = timezone.now()
                for detalle in pedido_locked.detalles.select_related('producto').all():
                    prod = Producto.objects.select_for_update().get(pk=detalle.producto.pk)
                    real_deducido = min(prod.stock, detalle.cantidad)
                    deficit = detalle.cantidad - real_deducido
                    prod.stock -= real_deducido
                    prod.save(update_fields=['stock'])
                    MovimientoInventario.objects.create(
                        producto=prod, tipo='salida',
                        cantidad=real_deducido,
                        nota=f'Pedido {pedido_locked.ticket}'
                    )
                    if deficit > 0:
                        # Stock insuficiente al entregar — registrar discrepancia (Bug #2)
                        MovimientoInventario.objects.create(
                            producto=prod, tipo='merma',
                            cantidad=deficit,
                            nota=f'Déficit en pedido {pedido_locked.ticket} — {deficit} unidad(es) sin stock'
                        )

            elif nuevo == 'cancelado' and pedido_locked.estado == 'entregado':
                # Anulación de entrega — restaurar stock (Bug #1)
                for detalle in pedido_locked.detalles.select_related('producto').all():
                    prod = Producto.objects.select_for_update().get(pk=detalle.producto.pk)
                    prod.stock += detalle.cantidad
                    prod.save(update_fields=['stock'])
                    MovimientoInventario.objects.create(
                        producto=prod, tipo='ajuste',
                        cantidad=detalle.cantidad,
                        nota=f'Devolución por anulación de pedido {pedido_locked.ticket}'
                    )

            pedido_locked.save(update_fields=['estado', 'fecha_entrega'])

        messages.success(request, f'Pedido {pedido.ticket} → {dict(Pedido.ESTADO_CHOICES).get(nuevo)}')
    return redirect(next_url)


# ══════════════════════════════════════════════════════════════════════════════
# PEDIDOS DOCENTES — redirige al panel unificado
# ══════════════════════════════════════════════════════════════════════════════

@admin_required
def pedidos_docentes(request):
    return redirect(reverse('app_admin:pedidos'))


@admin_required
def pedido_docente_detalle(request, pk):
    pedido = get_object_or_404(
        PedidoDocente.objects.select_related('docente__perfil__user')
                            .prefetch_related('detalles__producto'),
        pk=pk
    )
    return render(request, 'app_admin/pedido_docente_detalle.html', _ctx({'pedido': pedido}))


TRANSICIONES_VALIDAS_DOCENTE = {
    'pendiente':  ['preparando', 'cancelado'],
    'preparando': ['listo', 'pendiente', 'cancelado'],
    'listo':      ['entregado', 'pendiente'],
    'entregado':  [],
    'cancelado':  [],
}


@admin_required
def pedido_docente_estado(request, pk):
    pedido = get_object_or_404(PedidoDocente, pk=pk)
    next_url = request.POST.get('next', reverse('app_admin:pedidos'))
    if request.method == 'POST':
        nuevo = request.POST.get('estado')
        estados_validos = [e[0] for e in PedidoDocente.ESTADO_CHOICES]
        if nuevo not in estados_validos:
            messages.error(request, 'Estado no válido.')
            return redirect(next_url)

        if nuevo not in TRANSICIONES_VALIDAS_DOCENTE.get(pedido.estado, []):
            messages.error(
                request,
                f'No se puede cambiar de "{pedido.get_estado_display()}" a "{dict(PedidoDocente.ESTADO_CHOICES).get(nuevo)}".'
            )
            return redirect(next_url)

        with transaction.atomic():
            pedido_locked = PedidoDocente.objects.select_for_update().get(pk=pk)

            if nuevo not in TRANSICIONES_VALIDAS_DOCENTE.get(pedido_locked.estado, []):
                messages.error(request, 'Transición de estado inválida.')
                return redirect(next_url)

            pedido_locked.estado = nuevo
            if nuevo == 'entregado':
                for detalle in pedido_locked.detalles.select_related('producto').all():
                    prod = Producto.objects.select_for_update().get(pk=detalle.producto.pk)
                    prod.stock = max(0, prod.stock - detalle.cantidad)
                    prod.save(update_fields=['stock'])
                    MovimientoInventario.objects.create(
                        producto=prod, tipo='salida',
                        cantidad=detalle.cantidad,
                        nota=f'Pedido docente {pedido_locked.ticket}'
                    )

            pedido_locked.save(update_fields=['estado'])

        messages.success(request, f'Pedido {pedido.ticket} → {dict(PedidoDocente.ESTADO_CHOICES).get(nuevo)}')
    return redirect(next_url)


# ══════════════════════════════════════════════════════════════════════════════
# FACTURA
# ══════════════════════════════════════════════════════════════════════════════

# Datos de la cafetería — actualizar según la institución real
CAFETERIA_CONFIG = {
    'nombre':       'Cafetería Punto Asis',
    'nit':          '900.000.000-0',               # Actualizar con NIT real
    'direccion':    'Colegio San Francisco de Asis', # Actualizar
    'ciudad':       'Cali',                  # Actualizar
    'departamento': 'Valle del cauca',                 # Actualizar
    'telefono':     '+57 317 3055541',             # Actualizar
    'email':        'cafeteria@asis.edu.co',        # Actualizar
    'regimen':      'Régimen Simplificado',
    'ciiu':         '5629',
    'actividad':    'Otros servicios de expendio de alimentos',
    'resolucion':   'Habilitación mediante acto administrativo institucional',
}


@admin_required
def factura_vista(request, pk):
    pedido = get_object_or_404(
        Pedido.objects.select_related(
            'estudiante__perfil__user',
            'estudiante__padre__perfil__user',
        ).prefetch_related('detalles__producto__categoria'),
        pk=pk
    )

    nombre_estudiante = pedido.estudiante.perfil.user.get_full_name() or pedido.estudiante.perfil.user.username
    nombre_padre, telefono_padre = '', ''
    if pedido.estudiante.padre:
        nombre_padre = pedido.estudiante.padre.perfil.user.get_full_name() or pedido.estudiante.padre.perfil.user.username
        telefono_padre = pedido.estudiante.padre.perfil.telefono

    # Datos del responsable (admin logueado actualmente)
    responsable = None
    try:
        pa = request.user.perfil.perfil_admin
        responsable = {
            'nombre':    request.user.get_full_name() or request.user.username,
            'documento': pa.documento or '—',
            'cargo':     pa.cargo or 'Administrador de Cafetería',
        }
    except Exception:
        pass

    return render(request, 'app_admin/factura.html', _ctx({
        'pedido':            pedido,
        'config':            CAFETERIA_CONFIG,
        'nombre_estudiante': nombre_estudiante,
        'nombre_padre':      nombre_padre,
        'telefono_padre':    telefono_padre,
        'responsable':       responsable,
    }))


@admin_required
def factura_pdf(request, pk):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer, HRFlowable
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
        import io
    except ImportError:
        messages.error(request, 'reportlab no está instalado.')
        return redirect('app_admin:pedido_detalle', pk=pk)

    pedido = get_object_or_404(
        Pedido.objects.select_related(
            'estudiante__perfil__user',
            'estudiante__padre__perfil__user',
        ).prefetch_related('detalles__producto__categoria'),
        pk=pk
    )

    cfg = CAFETERIA_CONFIG
    nombre_estudiante = pedido.estudiante.perfil.user.get_full_name() or pedido.estudiante.perfil.user.username
    nombre_padre = ''
    telefono_padre = ''
    if pedido.estudiante.padre:
        nombre_padre   = pedido.estudiante.padre.perfil.user.get_full_name() or pedido.estudiante.padre.perfil.user.username
        telefono_padre = pedido.estudiante.padre.perfil.telefono

    responsable_nombre = ''
    responsable_doc = ''
    try:
        pa = request.user.perfil.perfil_admin
        responsable_nombre = request.user.get_full_name() or request.user.username
        responsable_doc    = pa.documento or '—'
    except Exception:
        pass

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=1.8*cm, rightMargin=1.8*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )

    # ── Colores de la marca
    VERDE_OSC = colors.HexColor('#0f2d1e')
    VERDE_MED = colors.HexColor('#16a34a')
    VERDE_PAL = colors.HexColor('#f0fdf4')
    GRIS_OSC  = colors.HexColor('#374151')
    GRIS_MED  = colors.HexColor('#6b7280')
    GRIS_PAL  = colors.HexColor('#f9fafb')
    BORDE     = colors.HexColor('#e5e7eb')

    # ── Estilos de párrafo
    styles = getSampleStyleSheet()
    def st(name, **kw):
        s = ParagraphStyle(name, **kw)
        return s

    bold_sm  = st('bsm',  fontName='Helvetica-Bold',   fontSize=7,  textColor=GRIS_OSC,  leading=10)
    norm_sm  = st('nsm',  fontName='Helvetica',         fontSize=7,  textColor=GRIS_MED,  leading=10)
    norm_xs  = st('nxs',  fontName='Helvetica',         fontSize=6,  textColor=GRIS_MED,  leading=9)
    head_lg  = st('hlg',  fontName='Helvetica-Bold',    fontSize=16, textColor=colors.white, leading=18)
    head_sm  = st('hsm',  fontName='Helvetica-Bold',    fontSize=8,  textColor=colors.white, leading=10)
    head_xs  = st('hxs',  fontName='Helvetica',         fontSize=6.5,textColor=colors.HexColor('#a3e8b8'), leading=9)
    num_lg   = st('nlg',  fontName='Helvetica-Bold',    fontSize=18, textColor=colors.HexColor('#7fffa0'), leading=20, alignment=TA_RIGHT)
    num_lbl  = st('nll',  fontName='Helvetica',         fontSize=6,  textColor=colors.HexColor('#a3e8b8'), leading=8,  alignment=TA_RIGHT)
    sec_lbl  = st('slab', fontName='Helvetica-Bold',    fontSize=6,  textColor=GRIS_MED,  leading=8, spaceAfter=2)
    sec_val  = st('sval', fontName='Helvetica-Bold',    fontSize=8.5,textColor=GRIS_OSC,  leading=11)
    tbl_hdr  = st('thdr', fontName='Helvetica-Bold',    fontSize=6.5,textColor=colors.white, leading=9, alignment=TA_CENTER)
    tbl_lft  = st('tlft', fontName='Helvetica',         fontSize=7.5,textColor=GRIS_OSC,  leading=10)
    tbl_rt   = st('trt',  fontName='Helvetica',         fontSize=7.5,textColor=GRIS_OSC,  leading=10, alignment=TA_RIGHT)
    tbl_bold = st('tbld', fontName='Helvetica-Bold',    fontSize=7.5,textColor=GRIS_OSC,  leading=10, alignment=TA_RIGHT)
    total_lbl= st('tolab',fontName='Helvetica-Bold',    fontSize=8,  textColor=GRIS_OSC,  leading=10)
    total_gr = st('togr', fontName='Helvetica-Bold',    fontSize=11, textColor=VERDE_OSC, leading=13, alignment=TA_RIGHT)
    foot_txt = st('ftxt', fontName='Helvetica',         fontSize=6,  textColor=GRIS_MED,  leading=8)
    leyenda  = st('ley',  fontName='Helvetica-Oblique', fontSize=5.5,textColor=GRIS_MED,  leading=8)

    W = doc.width  # ancho útil

    elements = []

    # ══════════════════════════════════════════════════════════════════
    # CABECERA — tabla de dos columnas con fondo verde oscuro
    # ══════════════════════════════════════════════════════════════════
    cab_emisor = [
        [Paragraph('CAFETERÍA ESCOLAR', head_xs)],
        [Paragraph('Punto Asis', head_lg)],
        [Paragraph(f"NIT: {cfg['nit']}", head_xs)],
        [Paragraph(cfg['direccion'],     head_xs)],
        [Paragraph(f"{cfg['ciudad']} · {cfg['departamento']}", head_xs)],
        [Paragraph(cfg['telefono'],      head_xs)],
        [Paragraph(cfg['email'],         head_xs)],
        [Paragraph(cfg['regimen'],       head_sm)],
    ]
    cab_doc = [
        [Paragraph('FACTURA DE VENTA', head_xs)],
        [Paragraph(pedido.ticket,       num_lg)],
        [Paragraph('N° de documento',   num_lbl)],
        [Spacer(1, 0.3*cm)],
        [Paragraph(f"Fecha: {pedido.fecha_pedido.strftime('%d/%m/%Y')}", head_xs)],
        [Paragraph(f"Hora:  {pedido.fecha_pedido.strftime('%H:%M')} hrs", head_xs)],
    ]

    cab_data = [[
        Table(cab_emisor, colWidths=[W*0.55], style=TableStyle([
            ('TOPPADDING',    (0,0),(-1,-1), 1.5),
            ('BOTTOMPADDING', (0,0),(-1,-1), 1.5),
            ('LEFTPADDING',   (0,0),(-1,-1), 0),
            ('RIGHTPADDING',  (0,0),(-1,-1), 0),
        ])),
        Table(cab_doc, colWidths=[W*0.4], style=TableStyle([
            ('TOPPADDING',    (0,0),(-1,-1), 1.5),
            ('BOTTOMPADDING', (0,0),(-1,-1), 1.5),
            ('LEFTPADDING',   (0,0),(-1,-1), 0),
            ('RIGHTPADDING',  (0,0),(-1,-1), 0),
            ('ALIGN',         (0,0),(-1,-1), 'RIGHT'),
        ])),
    ]]
    cab_table = Table(cab_data, colWidths=[W*0.6, W*0.4])
    cab_table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),(-1,-1), VERDE_OSC),
        ('TOPPADDING',    (0,0),(-1,-1), 14),
        ('BOTTOMPADDING', (0,0),(-1,-1), 14),
        ('LEFTPADDING',   (0,0),(-1,-1), 14),
        ('RIGHTPADDING',  (0,0),(-1,-1), 14),
        ('VALIGN',        (0,0),(-1,-1), 'TOP'),
        ('ROUNDEDCORNERS',(0,0),(-1,-1), [6,6,0,0]),
    ]))
    elements.append(cab_table)

    # ── Barra resolución DIAN
    res_data = [[Paragraph(
        f"Habilitación: {cfg['resolucion']} · CIIU {cfg['ciiu']} — {cfg['actividad']}",
        st('rb', fontName='Helvetica', fontSize=5.5, textColor=colors.HexColor('#a3e8b8'), leading=8)
    )]]
    res_table = Table(res_data, colWidths=[W])
    res_table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),(-1,-1), colors.HexColor('#193d2a')),
        ('TOPPADDING',    (0,0),(-1,-1), 4),
        ('BOTTOMPADDING', (0,0),(-1,-1), 4),
        ('LEFTPADDING',   (0,0),(-1,-1), 14),
        ('RIGHTPADDING',  (0,0),(-1,-1), 14),
    ]))
    elements.append(res_table)

    # ── Cinta de fechas
    ribbon_data = [[
        Table([[Paragraph('FECHA EXPEDICIÓN', sec_lbl)],[Paragraph(pedido.fecha_pedido.strftime('%d/%m/%Y'), sec_val)]], colWidths=[W/3]),
        Table([[Paragraph('FORMA DE PAGO',    sec_lbl)],[Paragraph('Saldo prepagado escolar',               sec_val)]], colWidths=[W/3]),
        Table([[Paragraph('ESTADO',           sec_lbl)],[Paragraph(pedido.get_estado_display(),              sec_val)]], colWidths=[W/3]),
    ]]
    ribbon_table = Table(ribbon_data, colWidths=[W/3, W/3, W/3])
    ribbon_table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),(-1,-1), GRIS_PAL),
        ('TOPPADDING',    (0,0),(-1,-1), 8),
        ('BOTTOMPADDING', (0,0),(-1,-1), 8),
        ('LEFTPADDING',   (0,0),(-1,-1), 12),
        ('RIGHTPADDING',  (0,0),(-1,-1), 12),
        ('LINEAFTER',     (0,0),(1,-1),  0.5, BORDE),
        ('LINEBELOW',     (0,0),(-1,-1), 0.5, BORDE),
    ]))
    elements.append(ribbon_table)

    # ══════════════════════════════════════════════════════════════════
    # PARTES — VENDEDOR / ADQUIRENTE
    # ══════════════════════════════════════════════════════════════════
    def parte_cell(titulo, rows):
        inner = [[Paragraph(titulo.upper(), st('ptag', fontName='Helvetica-Bold', fontSize=6, textColor=GRIS_MED, leading=8))]]
        for lbl, val in rows:
            if val:
                inner.append([Paragraph(f'<b>{lbl}:</b> {val}', norm_sm)])
        return Table(inner, colWidths=[(W/2 - 1*cm)], style=TableStyle([
            ('TOPPADDING',    (0,0),(-1,-1), 1.5),
            ('BOTTOMPADDING', (0,0),(-1,-1), 1.5),
            ('LEFTPADDING',   (0,0),(-1,-1), 0),
            ('RIGHTPADDING',  (0,0),(-1,-1), 0),
        ]))

    vendedor_rows = [
        ('NIT',        cfg['nit']),
        ('Dirección',  cfg['direccion']),
        ('Ciudad',     cfg['ciudad']),
        ('Régimen',    cfg['regimen']),
        ('Teléfono',   cfg['telefono']),
        ('Responsable',f"{responsable_nombre} — CC {responsable_doc}" if responsable_nombre else ''),
    ]
    adquirente_rows = [
        ('Nombre',        nombre_estudiante),
        ('Tipo doc.',     'Tarjeta de Identidad / Código estudiantil'),
        ('N° documento',  pedido.estudiante.codigo),
        ('Grado',         pedido.estudiante.grado),
        ('Acudiente',     nombre_padre),
        ('Tel. acudiente',telefono_padre),
        ('Ciudad',        cfg['ciudad']),
    ]

    partes_data = [[
        parte_cell('Vendedor / Emisor',     vendedor_rows),
        parte_cell('Adquirente / Comprador', adquirente_rows),
    ]]
    partes_table = Table(partes_data, colWidths=[W/2, W/2])
    partes_table.setStyle(TableStyle([
        ('TOPPADDING',    (0,0),(-1,-1), 10),
        ('BOTTOMPADDING', (0,0),(-1,-1), 10),
        ('LEFTPADDING',   (0,0),(-1,-1), 14),
        ('RIGHTPADDING',  (0,0),(-1,-1), 14),
        ('LINEAFTER',     (0,0),(0,-1),  0.5, BORDE),
        ('LINEBELOW',     (0,0),(-1,-1), 0.5, BORDE),
        ('VALIGN',        (0,0),(-1,-1), 'TOP'),
    ]))
    elements.append(partes_table)

    # ══════════════════════════════════════════════════════════════════
    # TABLA DE PRODUCTOS
    # ══════════════════════════════════════════════════════════════════
    prod_header = [
        Paragraph('DESCRIPCIÓN', tbl_hdr),
        Paragraph('CANT.', tbl_hdr),
        Paragraph('VLR. UNIT.', tbl_hdr),
        Paragraph('DESCUENTO', tbl_hdr),
        Paragraph('IVA', tbl_hdr),
        Paragraph('VLR. TOTAL', tbl_hdr),
    ]
    prod_rows = [prod_header]
    for d in pedido.detalles.all():
        prod_rows.append([
            Paragraph(f'<b>{d.producto.nombre}</b><br/><font size="6" color="#9ca3af">{d.producto.categoria} · {d.producto.get_tipo_display()}</font>', tbl_lft),
            Paragraph(str(d.cantidad),                 tbl_rt),
            Paragraph(f'${float(d.precio_unitario):,.0f}', tbl_rt),
            Paragraph('$0',                            tbl_rt),
            Paragraph('0% Excl.',                      tbl_rt),
            Paragraph(f'<b>${d.subtotal:,.0f}</b>',    tbl_bold),
        ])

    col_w = [W*0.38, W*0.08, W*0.14, W*0.12, W*0.12, W*0.16]
    prod_table = Table(prod_rows, colWidths=col_w, repeatRows=1)
    prod_table.setStyle(TableStyle([
        # Cabecera
        ('BACKGROUND',    (0,0),(-1,0),  VERDE_OSC),
        ('TEXTCOLOR',     (0,0),(-1,0),  colors.white),
        ('FONTNAME',      (0,0),(-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0,0),(-1,0),  6.5),
        ('TOPPADDING',    (0,0),(-1,0),  6),
        ('BOTTOMPADDING', (0,0),(-1,0),  6),
        ('LEFTPADDING',   (0,0),(-1,-1), 8),
        ('RIGHTPADDING',  (0,0),(-1,-1), 8),
        # Filas alternas
        ('ROWBACKGROUNDS', (0,1),(-1,-1), [colors.white, GRIS_PAL]),
        ('TOPPADDING',    (0,1),(-1,-1),  5),
        ('BOTTOMPADDING', (0,1),(-1,-1),  5),
        ('GRID',          (0,0),(-1,-1),  0.3, BORDE),
        ('VALIGN',        (0,0),(-1,-1),  'MIDDLE'),
        ('ALIGN',         (1,0),(-1,-1),  'RIGHT'),
        ('LINEBELOW',     (0,0),(-1,-1),  0.3, BORDE),
    ]))
    elements.append(prod_table)
    elements.append(Spacer(1, 0.4*cm))

    # ══════════════════════════════════════════════════════════════════
    # TOTALES + LEYENDA
    # ══════════════════════════════════════════════════════════════════
    total = float(pedido.total)

    def tot_row(lbl, val, bold=False):
        ls = st('trl', fontName='Helvetica-Bold' if bold else 'Helvetica', fontSize=8 if bold else 7.5, textColor=GRIS_OSC, leading=10)
        rs = st('trr', fontName='Helvetica-Bold' if bold else 'Helvetica', fontSize=8 if bold else 7.5, textColor=GRIS_OSC, leading=10, alignment=TA_RIGHT)
        return [Paragraph(lbl, ls), Paragraph(val, rs)]

    totales_data = [
        tot_row('Subtotal bruto',             f'${total:,.0f}'),
        tot_row('Descuentos',                 '$0'),
        tot_row('Base gravable',              f'${total:,.0f}'),
        tot_row('IVA (0% — Excluido)',        '$0'),
        tot_row('Retención en la fuente',     '$0'),
    ]

    # Fila TOTAL
    totales_data.append([
        Paragraph('TOTAL A PAGAR (COP)', st('trgd', fontName='Helvetica-Bold', fontSize=10, textColor=VERDE_OSC, leading=12)),
        Paragraph(f'${total:,.0f}',      st('trgv', fontName='Helvetica-Bold', fontSize=12, textColor=VERDE_OSC, leading=14, alignment=TA_RIGHT)),
    ])

    totales_table = Table(totales_data, colWidths=[W*0.5, W*0.22])
    totales_table.setStyle(TableStyle([
        ('TOPPADDING',    (0,0),(-1,-1), 3),
        ('BOTTOMPADDING', (0,0),(-1,-1), 3),
        ('LEFTPADDING',   (0,0),(-1,-1), 0),
        ('RIGHTPADDING',  (0,0),(-1,-1), 0),
        ('LINEBELOW',     (0,-2),(-1,-2), 0.5, BORDE),
        ('LINEABOVE',     (0,-1),(-1,-1), 1.5, VERDE_OSC),
        ('TOPPADDING',    (0,-1),(-1,-1), 6),
    ]))

    leyenda_data = [[Paragraph(
        'Los bienes y servicios de alimentación escolar están excluidos de IVA según el Art. 424 del '
        'Estatuto Tributario y el Decreto 1625 de 2016. No somos responsables del IVA (Régimen Simplificado). '
        'Este documento es válido como comprobante de pago para uso interno escolar.',
        leyenda
    )]]
    leyenda_table = Table(leyenda_data, colWidths=[W*0.5])
    leyenda_table.setStyle(TableStyle([
        ('TOPPADDING',    (0,0),(-1,-1), 6),
        ('BOTTOMPADDING', (0,0),(-1,-1), 4),
        ('LEFTPADDING',   (0,0),(-1,-1), 0),
        ('RIGHTPADDING',  (0,0),(-1,-1), 0),
    ]))

    pagado_data = [[Paragraph(
        '✓ PAGADO — Saldo prepagado descontado exitosamente',
        st('pg', fontName='Helvetica-Bold', fontSize=8, textColor=VERDE_MED, leading=10)
    )]]
    pagado_table = Table(pagado_data, colWidths=[W*0.5])
    pagado_table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),(-1,-1), VERDE_PAL),
        ('TOPPADDING',    (0,0),(-1,-1), 6),
        ('BOTTOMPADDING', (0,0),(-1,-1), 6),
        ('LEFTPADDING',   (0,0),(-1,-1), 8),
        ('RIGHTPADDING',  (0,0),(-1,-1), 8),
        ('ROUNDEDCORNERS',(0,0),(-1,-1), [4,4,4,4]),
    ]))

    # Observaciones del pedido (si las hay)
    if pedido.nota:
        obs_data = [[Paragraph(f'<b>Observaciones:</b> {pedido.nota}', norm_sm)]]
        obs_table = Table(obs_data, colWidths=[W*0.5])
        obs_table.setStyle(TableStyle([
            ('TOPPADDING',    (0,0),(-1,-1), 4),
            ('BOTTOMPADDING', (0,0),(-1,-1), 4),
            ('LEFTPADDING',   (0,0),(-1,-1), 0),
            ('RIGHTPADDING',  (0,0),(-1,-1), 0),
        ]))
    else:
        obs_table = Spacer(1, 0.2*cm)

    bottom_data = [[
        Table([
            [leyenda_table],
            [Spacer(1, 0.2*cm)],
            [obs_table],
            [Spacer(1, 0.2*cm)],
            [pagado_table],
        ], colWidths=[W*0.52]),
        Table([
            [totales_table],
        ], colWidths=[W*0.44]),
    ]]
    bottom = Table(bottom_data, colWidths=[W*0.54, W*0.46])
    bottom.setStyle(TableStyle([
        ('VALIGN',        (0,0),(-1,-1), 'BOTTOM'),
        ('TOPPADDING',    (0,0),(-1,-1), 0),
        ('BOTTOMPADDING', (0,0),(-1,-1), 0),
        ('LEFTPADDING',   (0,0),(-1,-1), 0),
        ('RIGHTPADDING',  (0,0),(-1,-1), 0),
    ]))
    elements.append(bottom)
    elements.append(Spacer(1, 0.4*cm))

    # ══════════════════════════════════════════════════════════════════
    # FOOTER
    # ══════════════════════════════════════════════════════════════════
    elements.append(HRFlowable(width=W, thickness=0.5, color=BORDE))
    elements.append(Spacer(1, 0.2*cm))

    foot_data = [[
        Paragraph(
            f'Punto Asis · {cfg["direccion"]}, {cfg["ciudad"]} · {cfg["telefono"]} · {cfg["email"]}',
            foot_txt
        ),
        Paragraph(
            f'<b>{pedido.ticket}</b> · {pedido.fecha_pedido.strftime("%d/%m/%Y %H:%M")}',
            st('ftr', fontName='Helvetica', fontSize=6, textColor=GRIS_MED, leading=8, alignment=TA_RIGHT)
        ),
    ]]
    foot_table = Table(foot_data, colWidths=[W*0.6, W*0.4])
    foot_table.setStyle(TableStyle([
        ('TOPPADDING',    (0,0),(-1,-1), 0),
        ('BOTTOMPADDING', (0,0),(-1,-1), 0),
        ('LEFTPADDING',   (0,0),(-1,-1), 0),
        ('RIGHTPADDING',  (0,0),(-1,-1), 0),
        ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
    ]))
    elements.append(foot_table)

    doc.build(elements)
    buffer.seek(0)

    nombre_archivo = f'factura_{pedido.ticket}.pdf'
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response


# ══════════════════════════════════════════════════════════════════════════════
# USUARIOS
# ══════════════════════════════════════════════════════════════════════════════

@admin_required
def usuarios(request):
    perfiles = Perfil.objects.select_related('user').order_by('rol', 'user__last_name')
    rol_f    = request.GET.get('rol', '')
    estado_f = request.GET.get('estado', '')
    q        = request.GET.get('q', '').strip()

    if rol_f:     perfiles = perfiles.filter(rol=rol_f)
    if estado_f == 'activo':   perfiles = perfiles.filter(activo=True)
    if estado_f == 'inactivo': perfiles = perfiles.filter(activo=False)
    if q:
        perfiles = perfiles.filter(
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q)  |
            Q(user__email__icontains=q)
        )

    conteos = {
        r['rol']: r['n']
        for r in Perfil.objects.values('rol').annotate(n=Count('id'))
    }
    activos = Perfil.objects.filter(activo=True).count()
    return render(request, 'app_admin/usuarios.html', _ctx({
        'perfiles':            perfiles,
        'rol_f':               rol_f,
        'estado_f':            estado_f,
        'q':                   q,
        'roles':               Perfil.ROL_CHOICES,
        'perfiles_activos':    activos,
        'perfiles_inactivos':  Perfil.objects.count() - activos,
        'perfiles_admin':      conteos.get('admin', 0),
        'perfiles_padre':      conteos.get('padre', 0),
        'perfiles_estudiante': conteos.get('estudiante', 0),
    }))


@admin_required
def usuario_toggle(request, pk):
    perfil = get_object_or_404(Perfil, pk=pk)
    if request.method == 'POST':
        perfil.activo = not perfil.activo
        perfil.save(update_fields=['activo'])
        messages.success(request, f'Usuario {perfil.user.username} {"activado" if perfil.activo else "desactivado"}.')
    return redirect('app_admin:usuarios')


@admin_required
def usuario_detalle(request, pk):
    perfil  = get_object_or_404(Perfil.objects.select_related('user'), pk=pk)
    context = _ctx({'perfil': perfil})
    if perfil.rol == 'estudiante':
        try:
            context['estudiante'] = perfil.estudiante
            context['pedidos']    = perfil.estudiante.pedidos.prefetch_related(
                'detalles__producto'
            ).order_by('-fecha_pedido')[:10]
        except: pass
    elif perfil.rol == 'padre':
        try:
            context['padre'] = perfil.padre
            context['hijos'] = perfil.padre.hijos.select_related('perfil__user').all()
        except: pass
    return render(request, 'app_admin/usuario_detalle.html', context)


@admin_required
def usuario_eliminar(request, pk):
    """Soft-delete: desactiva la cuenta de Django y el perfil."""
    perfil = get_object_or_404(Perfil, pk=pk)
    if request.method == 'POST':
        if perfil.rol == 'admin':
            messages.error(request, 'No puedes eliminar cuentas de administrador.')
            return redirect('app_admin:usuarios')
        username = perfil.user.username
        perfil.user.is_active = False
        perfil.user.save(update_fields=['is_active'])
        perfil.activo = False
        perfil.save(update_fields=['activo'])
        messages.success(request, f'Cuenta @{username} eliminada.')
    return redirect('app_admin:usuarios')


@admin_required
def usuario_bulk(request):
    """Acciones masivas sobre usuarios: activar, desactivar o eliminar."""
    if request.method == 'POST':
        pks    = request.POST.getlist('pks')
        accion = request.POST.get('accion', '')
        if not pks:
            messages.warning(request, 'No se seleccionó ningún usuario.')
            return redirect('app_admin:usuarios')
        # Solo afectar no-admins por seguridad
        perfiles = Perfil.objects.filter(pk__in=pks).exclude(rol='admin').select_related('user')
        count = perfiles.count()
        if accion == 'activar':
            perfiles.update(activo=True)
            messages.success(request, f'{count} usuario(s) activado(s).')
        elif accion == 'desactivar':
            perfiles.update(activo=False)
            messages.success(request, f'{count} usuario(s) desactivado(s).')
        elif accion == 'eliminar':
            for p in perfiles:
                p.user.is_active = False
                p.user.save(update_fields=['is_active'])
            perfiles.update(activo=False)
            messages.success(request, f'{count} usuario(s) eliminado(s).')
        else:
            messages.error(request, 'Acción inválida.')
    return redirect('app_admin:usuarios')


# ══════════════════════════════════════════════════════════════════════════════
# CARNETS QR (impresión)
# ══════════════════════════════════════════════════════════════════════════════

@admin_required
def carnets(request):
    """Lista de estudiantes o docentes para imprimir carnets QR."""
    tipo = request.GET.get('tipo', 'estudiante')
    q    = request.GET.get('q', '').strip()

    total_est = Estudiante.objects.filter(perfil__activo=True).count()
    total_doc = Docente.objects.filter(perfil__activo=True).count()

    if tipo == 'docente':
        qs = Docente.objects.select_related('perfil__user').filter(perfil__activo=True)
        if q:
            qs = qs.filter(
                Q(documento__icontains=q) |
                Q(perfil__user__first_name__icontains=q) |
                Q(perfil__user__last_name__icontains=q)
            )
        qs = qs.order_by('perfil__user__last_name', 'perfil__user__first_name')
        return render(request, 'app_admin/carnets.html', _ctx({
            'tipo':      'docente',
            'docentes':  qs,
            'q':         q,
            'total_est': total_est,
            'total_doc': qs.count(),
        }))

    # estudiante (default)
    grado = request.GET.get('grado', '').strip()
    qs = Estudiante.objects.select_related('perfil__user').filter(perfil__activo=True)
    if q:
        qs = qs.filter(
            Q(codigo__icontains=q) |
            Q(perfil__user__first_name__icontains=q) |
            Q(perfil__user__last_name__icontains=q)
        )
    if grado:
        qs = qs.filter(grado=grado)
    qs = qs.order_by('grado', 'perfil__user__last_name', 'perfil__user__first_name')
    grados_disponibles = (
        Estudiante.objects.filter(perfil__activo=True)
        .values_list('grado', flat=True).distinct().order_by('grado')
    )
    return render(request, 'app_admin/carnets.html', _ctx({
        'tipo':               'estudiante',
        'estudiantes':        qs,
        'q':                  q,
        'grado_f':            grado,
        'grados_disponibles': list(grados_disponibles),
        'total_est':          qs.count(),
        'total_doc':          total_doc,
    }))


@admin_required
def carnets_pdf(request):
    """Genera un PDF imprimible con los carnets seleccionados (8 por A4)."""
    import io
    import qrcode
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas as pdfcanvas
    from reportlab.lib.utils import ImageReader

    tipo = request.GET.get('tipo', 'estudiante')
    pks  = request.GET.getlist('pk') or request.POST.getlist('pk')
    if not pks:
        messages.warning(request, 'No seleccionaste ningún registro.')
        return redirect('app_admin:carnets')

    if tipo == 'docente':
        personas = list(
            Docente.objects
            .select_related('perfil__user')
            .filter(pk__in=pks, perfil__activo=True)
            .order_by('perfil__user__last_name', 'perfil__user__first_name')
        )
    else:
        personas = list(
            Estudiante.objects
            .select_related('perfil__user')
            .filter(pk__in=pks, perfil__activo=True)
            .order_by('grado', 'perfil__user__last_name', 'perfil__user__first_name')
        )

    buf = io.BytesIO()
    c = pdfcanvas.Canvas(buf, pagesize=A4)
    page_w, page_h = A4

    card_w   = 85.6 * mm
    card_h   = 54   * mm
    margin_x = (page_w - 2 * card_w) / 3
    margin_y = (page_h - 4 * card_h) / 5

    def dibujar_carnet(x, y, persona):
        c.setStrokeColorRGB(0.7, 0.7, 0.7)
        c.setLineWidth(0.5)
        c.roundRect(x, y, card_w, card_h, 6, stroke=1, fill=0)

        if tipo == 'docente':
            qr_data   = f'DOC-{persona.pk}'
            sub_label = persona.materia or 'Docente'
            id_label  = f'Doc: {persona.documento}' if persona.documento else ''
        else:
            qr_data   = persona.codigo
            sub_label = f'Grado: {persona.grado}'
            id_label  = f'Código: {persona.codigo}'

        qr = qrcode.QRCode(box_size=10, border=2)
        qr.add_data(qr_data)
        qr.make(fit=True)
        img     = qr.make_image(fill_color='black', back_color='white')
        img_buf = io.BytesIO()
        img.save(img_buf, format='PNG')
        img_buf.seek(0)
        qr_size = card_h - 10 * mm
        c.drawImage(ImageReader(img_buf), x + 4 * mm, y + 5 * mm, qr_size, qr_size)

        text_x      = x + qr_size + 8 * mm
        nombre      = persona.perfil.user.get_full_name() or persona.perfil.user.username
        nombre_corto = nombre if len(nombre) <= 28 else nombre[:26] + '…'

        c.setFillColorRGB(0.06, 0.06, 0.05)
        c.setFont('Helvetica-Bold', 7)
        c.drawString(text_x, y + card_h - 6 * mm, 'PUNTO ASIS')
        c.setFont('Helvetica', 6)
        c.setFillColorRGB(0.45, 0.45, 0.45)
        c.drawString(text_x, y + card_h - 9.5 * mm, 'Cafetería escolar')

        c.setFillColorRGB(0.06, 0.06, 0.05)
        c.setFont('Helvetica-Bold', 9)
        c.drawString(text_x, y + card_h - 18 * mm, nombre_corto)

        c.setFont('Helvetica', 7.5)
        c.setFillColorRGB(0.30, 0.30, 0.30)
        c.drawString(text_x, y + card_h - 23.5 * mm, sub_label)

        if id_label:
            c.setFont('Helvetica-Bold', 8)
            c.setFillColorRGB(0.06, 0.06, 0.05)
            c.drawString(text_x, y + card_h - 30 * mm, id_label)

        c.setFont('Helvetica', 5.5)
        c.setFillColorRGB(0.55, 0.55, 0.55)
        c.drawString(x + 4 * mm, y + 2 * mm, 'Presenta este carnet en caja para identificarte.')

    cards_per_page = 8
    for i, persona in enumerate(personas):
        slot = i % cards_per_page
        col  = slot % 2
        row  = slot // 2
        x    = margin_x + col * (card_w + margin_x)
        y    = page_h - margin_y - (row + 1) * card_h - row * margin_y
        dibujar_carnet(x, y, persona)
        if slot == cards_per_page - 1 and i + 1 < len(personas):
            c.showPage()

    c.save()
    buf.seek(0)

    fecha    = date.today().strftime('%Y%m%d')
    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="carnets_{tipo}_{fecha}.pdf"'
    return response


# ══════════════════════════════════════════════════════════════════════════════
# ESTADÍSTICAS
# ══════════════════════════════════════════════════════════════════════════════

@admin_required
def estadisticas(request):
    rango = request.GET.get('rango', '7')
    hoy   = date.today()
    desde = hoy - timedelta(days=int(rango))

    # Filtrar solo pedidos entregados para ingresos reales
    pedidos = Pedido.objects.filter(fecha_pedido__date__gte=desde, estado='entregado')

    agg      = pedidos.aggregate(t=Sum('total'), c=Sum('costo_total'))
    ingresos = float(agg['t'] or 0)
    costos   = float(agg['c'] or 0)
    ganancia  = ingresos - costos
    margen    = round((ganancia / ingresos) * 100, 1) if ingresos > 0 else 0
    n_pedidos = pedidos.count()

    top_productos = DetallePedido.objects.filter(pedido__in=pedidos).values(
        nombre=F('producto__nombre')
    ).annotate(
        unidades=Sum('cantidad'),
        ingresos=Sum(F('cantidad') * F('precio_unitario')),
        costos=Sum(F('cantidad') * F('costo_unitario')),
    ).order_by('-unidades')[:10]

    for p in top_productos:
        ingresos_val = float(p['ingresos'] or 0)
        costos_val = float(p['costos'] or 0)
        p['ganancia'] = round(ingresos_val - costos_val, 2)

    return render(request, 'app_admin/estadisticas.html', _ctx({
        'ingresos': ingresos, 'costos': costos, 'ganancia': ganancia,
        'margen': margen, 'n_pedidos': n_pedidos,
        'top_productos': top_productos, 'rango': rango, 'desde': desde,
    }))


@admin_required
def api_ventas(request):
    from django.core.cache import cache
    from django.db.models.functions import TruncDate

    try:
        dias = int(request.GET.get('dias', 7))
        if dias < 1 or dias > 365:
            dias = 7
    except (ValueError, TypeError):
        dias = 7

    cache_key = f'api_ventas_{date.today().isoformat()}_{dias}'
    resultado = cache.get(cache_key)
    if resultado is None:
        hoy   = date.today()
        desde = hoy - timedelta(days=dias - 1)
        rows  = {
            row['dia']: row
            for row in Pedido.objects.filter(
                fecha_pedido__date__gte=desde, estado='entregado'
            ).annotate(dia=TruncDate('fecha_pedido')).values('dia').annotate(
                ingresos=Sum('total'), costos=Sum('costo_total')
            )
        }
        labels, ingresos_data, ganancia_data = [], [], []
        for i in range(dias - 1, -1, -1):
            d   = hoy - timedelta(days=i)
            row = rows.get(d, {})
            ing = float(row.get('ingresos') or 0)
            cos = float(row.get('costos') or 0)
            labels.append(d.strftime('%d/%m'))
            ingresos_data.append(ing)
            ganancia_data.append(round(ing - cos, 2))
        resultado = {'labels': labels, 'ingresos': ingresos_data, 'ganancia': ganancia_data}
        cache.set(cache_key, resultado, timeout=1800)  # 30 min
    return JsonResponse(resultado)


@admin_required
def api_categorias(request):
    from django.core.cache import cache

    dias      = int(request.GET.get('dias', 7))
    cache_key = f'api_categorias_{date.today().isoformat()}_{dias}'
    resultado = cache.get(cache_key)
    if resultado is None:
        desde = date.today() - timedelta(days=dias)
        data  = DetallePedido.objects.filter(
            pedido__fecha_pedido__date__gte=desde, pedido__estado='entregado'
        ).values(cat=F('producto__categoria__nombre')).annotate(
            total=Sum(F('cantidad') * F('precio_unitario'))
        ).order_by('-total')
        resultado = {
            'labels': [d['cat'] for d in data],
            'values': [float(d['total'] or 0) for d in data],
        }
        cache.set(cache_key, resultado, timeout=1800)
    return JsonResponse(resultado)


@admin_required
def api_productos_top(request):
    from django.core.cache import cache

    dias      = int(request.GET.get('dias', 7))
    cache_key = f'api_productos_top_{date.today().isoformat()}_{dias}'
    resultado = cache.get(cache_key)
    if resultado is None:
        desde = date.today() - timedelta(days=dias)
        data  = DetallePedido.objects.filter(
            pedido__fecha_pedido__date__gte=desde, pedido__estado='entregado'
        ).values(nombre=F('producto__nombre')).annotate(
            unidades=Sum('cantidad')
        ).order_by('-unidades')[:8]
        resultado = {
            'labels': [d['nombre'] for d in data],
            'values': [d['unidades'] for d in data],
        }
        cache.set(cache_key, resultado, timeout=1800)
    return JsonResponse(resultado)


# ══════════════════════════════════════════════════════════════════════════════
# EXPORTAR
# ══════════════════════════════════════════════════════════════════════════════

@admin_required
def exportar_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messages.error(request, 'openpyxl no está instalado. Ejecuta: pip install openpyxl')
        return redirect('app_admin:inventario')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventario'

    # Estilo cabecera
    header_fill = PatternFill(start_color='1a3d2b', end_color='1a3d2b', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    headers = ['Nombre', 'Categoría', 'Tipo', 'Proveedor',
               'Precio Costo', 'Precio Venta', 'Ganancia', 'Margen %', 'Stock', 'Stock Mínimo', 'Disponible']
    ws.append(headers)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for p in Producto.objects.select_related('categoria', 'proveedor').all():
        ws.append([
            p.nombre,
            str(p.categoria),
            p.get_tipo_display(),
            str(p.proveedor) if p.proveedor else '—',
            float(p.precio_costo or 0),
            float(p.precio_venta),
            p.ganancia,
            p.margen,
            p.stock if p.tipo == 'simple' else '—',
            p.stock_minimo if p.tipo == 'simple' else '—',
            'Sí' if p.disponible else 'No',
        ])

    # Ajustar anchos
    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="inventario_punto_asis.xlsx"'
    wb.save(response)
    return response


@admin_required
def exportar_pdf(request):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        import io
    except ImportError:
        messages.error(request, 'reportlab no está instalado. Ejecuta: pip install reportlab')
        return redirect('app_admin:inventario')

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = []

    # Título
    elements.append(Paragraph('Inventario — Punto Asis', styles['Title']))
    elements.append(Paragraph(f'Generado: {date.today().strftime("%d/%m/%Y")}', styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))

    # Tabla
    headers = ['Nombre', 'Categoría', 'Tipo', 'Proveedor',
               'P. Costo', 'P. Venta', 'Ganancia', 'Margen%', 'Stock']
    data = [headers]
    for p in Producto.objects.select_related('categoria', 'proveedor').all():
        data.append([
            p.nombre[:25],
            str(p.categoria),
            p.get_tipo_display()[:10],
            str(p.proveedor)[:15] if p.proveedor else '—',
            f'${float(p.precio_costo or 0):,.0f}',
            f'${float(p.precio_venta):,.0f}',
            f'${p.ganancia:,.0f}',
            f'{p.margen}%',
            str(p.stock) if p.tipo == 'simple' else '—',
        ])

    verde_oscuro = colors.HexColor('#1a3d2b')
    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), verde_oscuro),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f2f2ef')]),
        ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#e2e2dd')),
        ('ALIGN',      (4, 0), (-1, -1), 'RIGHT'),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(tabla)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="inventario_punto_asis.pdf"'
    return response


# ══════════════════════════════════════════════════════════════════════════════
# PERFIL Y CONFIGURACIÓN
# ══════════════════════════════════════════════════════════════════════════════

@admin_required
def perfil(request):
    perfil_auth = request.user.perfil
    perfil_admin, _ = PerfilAdmin.objects.get_or_create(perfil=perfil_auth)

    if request.method == 'POST':
        u = request.user
        u.first_name = request.POST.get('first_name', u.first_name).strip()
        u.last_name  = request.POST.get('last_name',  u.last_name).strip()
        u.email      = request.POST.get('email', u.email).strip()
        u.save(update_fields=['first_name', 'last_name', 'email'])

        perfil_auth.telefono = request.POST.get('telefono', '').strip()
        perfil_auth.save(update_fields=['telefono'])

        perfil_admin.documento        = request.POST.get('documento', '').strip()
        perfil_admin.direccion        = request.POST.get('direccion', '').strip()
        perfil_admin.cargo            = request.POST.get('cargo', '').strip()
        fecha_nac = request.POST.get('fecha_nacimiento') or None
        perfil_admin.fecha_nacimiento = fecha_nac
        if request.FILES.get('foto'):
            perfil_admin.foto = request.FILES['foto']
        perfil_admin.save()

        messages.success(request, 'Perfil actualizado correctamente.')
        return redirect('app_admin:perfil')

    return render(request, 'app_admin/perfil.html', _ctx({
        'perfil_auth':  perfil_auth,
        'perfil_admin': perfil_admin,
    }))


@admin_required
def configuracion(request):
    return render(request, 'app_admin/configuracion.html', _ctx())


# ══════════════════════════════════════════════════════════════════════════════
# ALERTAS API (para topbar)
# ══════════════════════════════════════════════════════════════════════════════

@admin_required
def api_alertas(request):
    alertas = _get_alertas()

    detalle = []

    for p in Producto.objects.filter(tipo='simple', stock__lte=F('stock_minimo')).select_related('proveedor')[:5]:
        detalle.append({
            'tipo': 'stock_producto',
            'texto': f'{p.nombre} — {p.stock} uds. (mín: {p.stock_minimo})',
            'url': f'/admin-panel/inventario/',
            'color': 'amber',
        })

    for i in [x for x in Ingrediente.objects.filter(activo=True) if x.stock_bajo][:5]:
        detalle.append({
            'tipo': 'stock_ingrediente',
            'texto': f'{i.nombre} — {i.stock_real:.2g} {i.get_unidad_base_display()} (mín: {i.stock_minimo})',
            'url': '/admin-panel/inventario/ingredientes/',
            'color': 'amber',
        })

    for lote in LoteIngrediente.objects.filter(
        ingrediente__activo=True,
        cantidad_base__gt=0,
        fecha_vencimiento__lt=date.today()
    ).select_related('ingrediente').order_by('fecha_vencimiento')[:3]:
        detalle.append({
            'tipo': 'vencido',
            'texto': f'{lote.ingrediente.nombre} venció el {lote.fecha_vencimiento.strftime("%d/%m/%Y")}',
            'url': '/admin-panel/inventario/ingredientes/',
            'color': 'red',
        })

    for lote in LoteIngrediente.objects.filter(
        ingrediente__activo=True,
        cantidad_base__gt=0,
        fecha_vencimiento__gte=date.today(),
        fecha_vencimiento__lte=date.today() + timedelta(days=7)
    ).select_related('ingrediente').order_by('fecha_vencimiento')[:3]:
        detalle.append({
            'tipo': 'vence_pronto',
            'texto': f'{lote.ingrediente.nombre} vence el {lote.fecha_vencimiento.strftime("%d/%m/%Y")}',
            'url': '/admin-panel/inventario/ingredientes/',
            'color': 'amber',
        })

    alertas['detalle'] = detalle
    return JsonResponse(alertas)


# ══════════════════════════════════════════════════════════════════════════════
# IA — GENERADOR DE DESCRIPCIÓN DE PRODUCTO
# ══════════════════════════════════════════════════════════════════════════════

@admin_required
@require_POST
def api_generar_descripcion(request):
    """AJAX: recibe nombre + categoría + ingredientes y devuelve descripción generada por Claude."""
    from edupos.ai_client import get_client
    import anthropic as _anthropic

    nombre      = request.POST.get('nombre', '').strip()
    categoria   = request.POST.get('categoria', '').strip()
    ingredientes = request.POST.getlist('ingredientes[]')

    if not nombre:
        return JsonResponse({'error': 'El nombre del producto es obligatorio.'}, status=400)

    cat_label = dict(Categoria.NOMBRE_CHOICES).get(categoria, categoria) or 'cafetería'
    ings_str  = ', '.join(ingredientes[:6]) if ingredientes else 'ingredientes frescos'

    prompt = (
        f'Eres el redactor de menú de una cafetería escolar colombiana llamada "Punto Asis".\n'
        f'Escribe una descripción apetitosa y corta (máximo 25 palabras) para este producto.\n'
        f'Usa lenguaje cercano y fresco, apto para niños y jóvenes. Sin comillas ni punto final.\n\n'
        f'Producto: "{nombre}"\n'
        f'Categoría: "{cat_label}"\n'
        f'Ingredientes principales: "{ings_str}"\n\n'
        f'Responde SOLO con la descripción.'
    )

    try:
        msg = get_client().messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=80,
            messages=[{'role': 'user', 'content': prompt}],
        )
        return JsonResponse({'descripcion': msg.content[0].text.strip()})
    except RuntimeError as e:
        return JsonResponse({'error': str(e)}, status=503)
    except _anthropic.APIError:
        return JsonResponse({'error': 'Servicio de IA no disponible en este momento.'}, status=503)


# ══════════════════════════════════════════════════════════════════════════════
# IA — EXTRACTOR DE INFORMACIÓN DE PAGO EN RECARGAS
# ══════════════════════════════════════════════════════════════════════════════

@admin_required
def api_info_pago(request, pk):
    """AJAX GET: analiza la nota libre del padre y extrae datos de pago estructurados."""
    from app_padre.models import RecargaSaldo
    from edupos.ai_client import get_client
    import anthropic as _anthropic

    recarga = get_object_or_404(RecargaSaldo, pk=pk)

    if not recarga.nota or len(recarga.nota.strip()) < 4:
        return JsonResponse({'sin_nota': True})

    prompt = (
        f'Extrae información de pago de esta nota escrita por un padre de familia colombiano.\n'
        f'La nota acompaña a una solicitud de recarga de saldo en una cafetería escolar.\n\n'
        f'Nota del padre: "{recarga.nota}"\n'
        f'Monto solicitado: ${float(recarga.monto):,.0f} COP\n\n'
        f'Responde SOLO con JSON válido:\n'
        f'{{\n'
        f'  "banco": "nombre del banco o null",\n'
        f'  "numero_referencia": "número de transacción/referencia o null",\n'
        f'  "medio_pago": "transferencia|nequi|daviplata|efectivo|otro|no_especificado",\n'
        f'  "fecha_mencionada": "fecha en texto o null",\n'
        f'  "resumen": "frase de máximo 10 palabras resumiendo el pago"\n'
        f'}}'
    )

    try:
        msg = get_client().messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=150,
            messages=[{'role': 'user', 'content': prompt}],
        )
        import json as _json
        data = _json.loads(msg.content[0].text.strip())
        return JsonResponse(data)
    except RuntimeError as e:
        return JsonResponse({'error': str(e)}, status=503)
    except (_anthropic.APIError, ValueError):
        return JsonResponse({'error': 'No disponible'}, status=503)


# ══════════════════════════════════════════════════════════════════════════════
# GESTIÓN DE RECARGAS (validación de comprobantes)
# ══════════════════════════════════════════════════════════════════════════════

@admin_required
def recargas(request):
    import types as _types
    from app_padre.models import RecargaSaldo, RecargaPadre
    from app_docente.models import RecargaDocente

    estado = request.GET.get('estado', 'pendiente')

    qs_est = RecargaSaldo.objects.select_related('estudiante__perfil__user', 'padre__perfil__user')
    qs_doc = RecargaDocente.objects.select_related('docente__perfil__user')
    qs_pad = RecargaPadre.objects.select_related('padre__perfil__user')
    if estado:
        qs_est = qs_est.filter(estado=estado)
        qs_doc = qs_doc.filter(estado=estado)
        qs_pad = qs_pad.filter(estado=estado)

    lista = []
    for r in qs_est:
        lista.append(_types.SimpleNamespace(
            pk=r.pk, tipo='estudiante',
            nombre=r.estudiante.perfil.user.get_full_name(),
            subtitulo=f'Padre: {r.padre.perfil.user.get_full_name() if r.padre else "–"}',
            rol_badge='Estudiante',
            monto=r.monto, comprobante=r.comprobante,
            nota=r.nota, nota_admin=r.nota_admin,
            estado=r.estado, get_estado_display=r.get_estado_display,
            fecha=r.fecha, fecha_resolucion=r.fecha_resolucion,
        ))
    for r in qs_doc:
        lista.append(_types.SimpleNamespace(
            pk=r.pk, tipo='docente',
            nombre=r.docente.perfil.user.get_full_name(),
            subtitulo='Docente — saldo propio',
            rol_badge='Docente',
            monto=r.monto, comprobante=r.comprobante,
            nota=r.nota, nota_admin=r.nota_admin,
            estado=r.estado, get_estado_display=r.get_estado_display,
            fecha=r.fecha, fecha_resolucion=r.fecha_resolucion,
        ))
    for r in qs_pad:
        lista.append(_types.SimpleNamespace(
            pk=r.pk, tipo='padre',
            nombre=r.padre.perfil.user.get_full_name(),
            subtitulo='Padre — saldo propio',
            rol_badge='Padre',
            monto=r.monto, comprobante=r.comprobante,
            nota=r.nota, nota_admin=r.nota_admin,
            estado=r.estado, get_estado_display=r.get_estado_display,
            fecha=r.fecha, fecha_resolucion=r.fecha_resolucion,
        ))

    lista.sort(key=lambda r: r.fecha, reverse=True)

    n_pendientes = (
        RecargaSaldo.objects.filter(estado='pendiente').count()
        + RecargaDocente.objects.filter(estado='pendiente').count()
        + RecargaPadre.objects.filter(estado='pendiente').count()
    )
    return render(request, 'app_admin/recargas.html', _ctx({
        'recargas': lista,
        'estado_activo': estado,
        'n_pendientes': n_pendientes,
    }))


@admin_required
def recarga_resolver(request, tipo, pk):
    from app_padre.models import RecargaSaldo, RecargaPadre
    from app_docente.models import RecargaDocente

    _MODELS = {'estudiante': RecargaSaldo, 'docente': RecargaDocente, 'padre': RecargaPadre}
    Model = _MODELS.get(tipo)
    if not Model:
        messages.error(request, 'Tipo de recarga inválido.')
        return redirect('app_admin:recargas')

    recarga = get_object_or_404(Model, pk=pk, estado='pendiente')
    if request.method == 'POST':
        accion     = request.POST.get('accion', '')
        nota_admin = request.POST.get('nota_admin', '').strip()
        if accion == 'aprobar':
            recarga.aprobar()
            messages.success(request, f'Recarga #{recarga.pk} de ${recarga.monto:,.0f} aprobada.')
        elif accion == 'rechazar':
            recarga.rechazar(nota=nota_admin)
            messages.warning(request, f'Recarga #{recarga.pk} rechazada.')
        else:
            messages.error(request, 'Acción inválida.')
    return redirect('app_admin:recargas')


# ══════════════════════════════════════════════════════════════════════════════
# INSUMOS
# ══════════════════════════════════════════════════════════════════════════════

@admin_required
def insumos(request):
    cat_f  = request.GET.get('cat', '')
    q      = request.GET.get('q', '').strip()
    qs     = Insumo.objects.select_related('proveedor').order_by('categoria', 'nombre')
    if cat_f:
        qs = qs.filter(categoria=cat_f)
    if q:
        qs = qs.filter(nombre__icontains=q)

    n_stock_bajo   = Insumo.objects.filter(activo=True, stock__lte=F('stock_minimo')).count()
    n_sin_stock    = Insumo.objects.filter(activo=True, stock__lte=0).count()
    n_con_proveedor = Insumo.objects.filter(activo=True, proveedor__isnull=False).count()
    movimientos    = MovimientoInsumo.objects.select_related('insumo').order_by('-fecha')[:20]

    return render(request, 'app_admin/insumos.html', _ctx({
        'insumos': qs,
        'cat_f': cat_f,
        'q': q,
        'CATEGORIA_CHOICES': Insumo.CATEGORIA_CHOICES,
        'n_stock_bajo': n_stock_bajo,
        'n_sin_stock': n_sin_stock,
        'n_con_proveedor': n_con_proveedor,
        'movimientos': movimientos,
    }))


@admin_required
def insumo_nuevo(request):
    if request.method == 'POST':
        nombre   = request.POST.get('nombre', '').strip()
        categoria = request.POST.get('categoria', 'otro')
        unidad   = request.POST.get('unidad', 'und')
        stock    = request.POST.get('stock', '0').strip()
        stock_min = request.POST.get('stock_minimo', '0').strip()
        precio   = request.POST.get('precio_unitario', '0').strip()
        prov_id  = request.POST.get('proveedor', '') or None
        desc     = request.POST.get('descripcion', '').strip()

        if not nombre:
            messages.error(request, 'El nombre es obligatorio.')
            return redirect('app_admin:insumo_nuevo')

        try:
            stock_f    = float(stock)
            stock_min_f = float(stock_min)
            precio_f   = float(precio)
            if stock_f < 0 or stock_min_f < 0 or precio_f < 0:
                raise ValueError
        except (ValueError, TypeError):
            messages.error(request, 'Stock y precio deben ser números positivos.')
            return redirect('app_admin:insumo_nuevo')

        imagen = request.FILES.get('imagen')
        insumo = Insumo.objects.create(
            nombre=nombre, categoria=categoria, unidad=unidad,
            stock=stock_f, stock_minimo=stock_min_f,
            precio_unitario=precio_f, proveedor_id=prov_id,
            descripcion=desc, imagen=imagen,
        )
        if stock_f > 0:
            MovimientoInsumo.objects.create(
                insumo=insumo, tipo='entrada',
                cantidad=stock_f, nota='Stock inicial'
            )
        messages.success(request, f'Insumo "{insumo.nombre}" creado exitosamente.')
        return redirect('app_admin:insumos')

    proveedores = Proveedor.objects.filter(activo=True)
    return render(request, 'app_admin/insumo_form.html', _ctx({
        'accion': 'Nuevo',
        'proveedores': proveedores,
        'CATEGORIA_CHOICES': Insumo.CATEGORIA_CHOICES,
        'UNIDAD_CHOICES': Insumo.UNIDAD_CHOICES,
    }))


@admin_required
def insumo_editar(request, pk):
    insumo = get_object_or_404(Insumo, pk=pk)
    if request.method == 'POST':
        insumo.nombre      = request.POST.get('nombre', '').strip() or insumo.nombre
        insumo.categoria   = request.POST.get('categoria', insumo.categoria)
        insumo.unidad      = request.POST.get('unidad', insumo.unidad)
        insumo.descripcion = request.POST.get('descripcion', '').strip()
        prov_id            = request.POST.get('proveedor', '') or None
        insumo.proveedor_id = prov_id

        try:
            insumo.stock_minimo    = float(request.POST.get('stock_minimo', insumo.stock_minimo))
            insumo.precio_unitario = float(request.POST.get('precio_unitario', insumo.precio_unitario))
        except (ValueError, TypeError):
            messages.error(request, 'Valores numéricos inválidos.')
            return redirect('app_admin:insumo_editar', pk=pk)

        if request.FILES.get('imagen'):
            insumo.imagen = request.FILES['imagen']
        insumo.save()
        messages.success(request, f'Insumo "{insumo.nombre}" actualizado.')
        return redirect('app_admin:insumos')

    proveedores = Proveedor.objects.filter(activo=True)
    return render(request, 'app_admin/insumo_form.html', _ctx({
        'accion': 'Editar',
        'insumo': insumo,
        'proveedores': proveedores,
        'CATEGORIA_CHOICES': Insumo.CATEGORIA_CHOICES,
        'UNIDAD_CHOICES': Insumo.UNIDAD_CHOICES,
    }))


@admin_required
def insumo_ajuste(request, pk):
    insumo = get_object_or_404(Insumo, pk=pk)
    if request.method == 'POST':
        tipo     = request.POST.get('tipo', 'entrada')
        nota     = request.POST.get('nota', '').strip()

        if tipo not in ('entrada', 'salida', 'ajuste', 'merma'):
            messages.error(request, 'Tipo de movimiento inválido.')
            return redirect('app_admin:insumos')

        try:
            cantidad = float(request.POST.get('cantidad', 0))
            if cantidad <= 0:
                raise ValueError
        except (ValueError, TypeError):
            messages.error(request, 'La cantidad debe ser mayor a 0.')
            return redirect('app_admin:insumos')

        with transaction.atomic():
            ins = Insumo.objects.select_for_update().get(pk=insumo.pk)
            if tipo in ('salida', 'merma'):
                ins.stock = max(0, float(ins.stock) - cantidad)
            elif tipo == 'ajuste':
                ins.stock = cantidad
            else:
                ins.stock = float(ins.stock) + cantidad
            ins.save(update_fields=['stock'])

            MovimientoInsumo.objects.create(
                insumo=ins, tipo=tipo, cantidad=cantidad, nota=nota
            )

        messages.success(request, f'Movimiento registrado para "{insumo.nombre}".')
        return redirect('app_admin:insumos')

    tipo_inicial = request.GET.get('tipo', 'entrada')
    return render(request, 'app_admin/insumo_ajuste.html', _ctx({
        'insumo':       insumo,
        'TIPO_CHOICES': MovimientoInsumo.TIPO_CHOICES,
        'tipo_inicial': tipo_inicial,
    }))


@admin_required
def insumo_toggle(request, pk):
    insumo = get_object_or_404(Insumo, pk=pk)
    if request.method == 'POST':
        insumo.activo = not insumo.activo
        insumo.save(update_fields=['activo'])
        estado = 'activado' if insumo.activo else 'desactivado'
        messages.success(request, f'Insumo "{insumo.nombre}" {estado}.')
    return redirect('app_admin:insumos')


@admin_required
def insumo_historial(request, pk):
    insumo      = get_object_or_404(Insumo, pk=pk)
    movimientos = insumo.movimientos.order_by('-fecha')
    totales = {
        'entrada': sum(float(m.cantidad) for m in movimientos if m.tipo == 'entrada'),
        'salida':  sum(float(m.cantidad) for m in movimientos if m.tipo == 'salida'),
        'merma':   sum(float(m.cantidad) for m in movimientos if m.tipo == 'merma'),
        'ajuste':  movimientos.filter(tipo='ajuste').count(),
    }
    return render(request, 'app_admin/insumo_historial.html', _ctx({
        'insumo':      insumo,
        'movimientos': movimientos,
        'totales':     totales,
    }))


@admin_required
def insumo_eliminar(request, pk):
    """Soft-delete de insumo."""
    insumo = get_object_or_404(Insumo, pk=pk)
    if request.method == 'POST':
        nombre = insumo.nombre
        insumo.activo = False
        insumo.save(update_fields=['activo'])
        messages.success(request, f'Insumo "{nombre}" eliminado del sistema.')
    return redirect('app_admin:insumos')


@admin_required
def exportar_insumos_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messages.error(request, 'openpyxl no está instalado.')
        return redirect('app_admin:insumos')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Insumos'

    header_fill = PatternFill(start_color='0a3055', end_color='0a3055', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    headers = ['Nombre', 'Categoría', 'Unidad', 'Stock actual', 'Stock mínimo',
               'Precio unitario', 'Proveedor', 'Descripción', 'Activo']
    ws.append(headers)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for ins in Insumo.objects.select_related('proveedor').order_by('categoria', 'nombre'):
        ws.append([
            ins.nombre,
            ins.get_categoria_display(),
            ins.get_unidad_display(),
            float(ins.stock),
            float(ins.stock_minimo),
            float(ins.precio_unitario),
            ins.proveedor.nombre if ins.proveedor else '—',
            ins.descripcion or '',
            'Sí' if ins.activo else 'No',
        ])

    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="insumos_punto_asis.xlsx"'
    wb.save(response)
    return response


@admin_required
def exportar_insumos_pdf(request):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        import io
    except ImportError:
        messages.error(request, 'reportlab no está instalado.')
        return redirect('app_admin:insumos')

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph('Insumos — Punto Asis', styles['Title']))
    elements.append(Paragraph(f'Generado: {date.today().strftime("%d/%m/%Y")}', styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))

    headers = ['Nombre', 'Categoría', 'Unidad', 'Stock', 'Mínimo', 'Precio/u', 'Proveedor', 'Estado']
    data = [headers]
    for ins in Insumo.objects.select_related('proveedor').order_by('categoria', 'nombre'):
        data.append([
            ins.nombre[:30],
            ins.get_categoria_display(),
            ins.get_unidad_display(),
            str(float(ins.stock)),
            str(float(ins.stock_minimo)),
            f'${float(ins.precio_unitario):,.2f}',
            ins.proveedor.nombre[:18] if ins.proveedor else '—',
            'Activo' if ins.activo else 'Inactivo',
        ])

    azul_oscuro = colors.HexColor('#0a3055')
    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), azul_oscuro),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f8ff')]),
        ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#e2e2dd')),
        ('ALIGN',      (3, 0), (5, -1), 'RIGHT'),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(tabla)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="insumos_punto_asis.pdf"'
    return response



# ══════════════════════════════════════════════════════════════════════════════
# EMPLEADOS
# ══════════════════════════════════════════════════════════════════════════════

import secrets
import string as _string


def _generar_password(length=10):
    alphabet = _string.ascii_letters + _string.digits
    return ''.join(secrets.choice(alphabet) for _ in range(length))


@admin_required
def empleados(request):
    from authentication.models import Empleado, Sede

    qs = Empleado.objects.select_related('perfil__user').prefetch_related('sedes').order_by('perfil__user__last_name')

    q       = request.GET.get('q', '').strip()
    sede_id = request.GET.get('sede', '')

    if q:
        qs = qs.filter(
            Q(perfil__user__first_name__icontains=q) |
            Q(perfil__user__last_name__icontains=q) |
            Q(perfil__user__email__icontains=q) |
            Q(documento__icontains=q)
        )
    if sede_id:
        qs = qs.filter(sedes__pk=sede_id)

    sedes = Sede.objects.filter(activa=True)

    all_emp = Empleado.objects.all()
    return render(request, 'app_admin/empleados.html', _ctx({
        'empleados':          qs,
        'sedes':              sedes,
        'q':                  q,
        'sede_id':            sede_id,
        'total_empleados':    all_emp.count(),
        'empleados_activos':  all_emp.filter(perfil__activo=True).count(),
        'empleados_inactivos':all_emp.filter(perfil__activo=False).count(),
    }))


@admin_required
def empleado_nuevo(request):
    from authentication.models import Empleado, Sede, Perfil
    from authentication.utils import generar_username
    from django.contrib.auth.models import User

    sedes = Sede.objects.filter(activa=True)

    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        email      = request.POST.get('email', '').strip()
        telefono   = request.POST.get('telefono', '').strip()
        documento  = request.POST.get('documento', '').strip()
        cargo      = request.POST.get('cargo', 'cajero')
        es_global  = request.POST.get('es_global') == 'on'
        sede_ids   = request.POST.getlist('sedes')

        password   = request.POST.get('password', '').strip()
        password2  = request.POST.get('password2', '').strip()

        errores = []
        if not first_name or not last_name:
            errores.append('Nombre y apellido son obligatorios.')
        if not password:
            errores.append('La contraseña es obligatoria.')
        elif len(password) < 6:
            errores.append('La contraseña debe tener al menos 6 caracteres.')
        elif password != password2:
            errores.append('Las contraseñas no coinciden.')
        if email and User.objects.filter(email=email).exists():
            errores.append('Ya existe una cuenta registrada con ese correo.')

        if errores:
            for e in errores:
                messages.error(request, e)
        else:
            from authentication.utils import _normalizar
            if email:
                username = generar_username(email)
            else:
                username = f"{_normalizar(first_name)}.{_normalizar(last_name)}@interno.local"
                email = username

            user = User.objects.create_user(
                username=username, password=password,
                first_name=first_name, last_name=last_name, email=email,
            )
            perfil   = Perfil.objects.create(user=user, rol='empleado', telefono=telefono)
            empleado = Empleado.objects.create(
                perfil=perfil, documento=documento, cargo=cargo, es_global=es_global,
            )
            if sede_ids:
                from authentication.models import Sede as SedeModel
                empleado.sedes.set(SedeModel.objects.filter(pk__in=sede_ids))

            messages.success(
                request,
                f'Empleado creado. Correo: <strong>{email}</strong>. '
                f'Credenciales listas para compartir.'
            )
            return redirect('app_admin:empleado_detalle', pk=empleado.pk)

    from authentication.models import Empleado as EmpModel
    return render(request, 'app_admin/empleado_nuevo.html', _ctx({
        'sedes':  sedes,
        'cargos': EmpModel.CARGO_CHOICES,
    }))


@admin_required
def empleado_detalle(request, pk):
    from authentication.models import Empleado, Sede

    empleado = get_object_or_404(Empleado, pk=pk)
    sedes    = Sede.objects.all()

    if request.method == 'POST':
        accion = request.POST.get('accion', 'editar')

        if accion == 'editar':
            empleado.documento = request.POST.get('documento', '').strip()
            empleado.cargo     = request.POST.get('cargo', 'cajero')
            empleado.es_global = request.POST.get('es_global') == 'on'
            sede_ids           = request.POST.getlist('sedes')
            empleado.sedes.set(Sede.objects.filter(pk__in=sede_ids))
            empleado.save()

            perfil = empleado.perfil
            perfil.telefono = request.POST.get('telefono', '').strip()
            perfil.save(update_fields=['telefono'])

            user = perfil.user
            user.first_name = request.POST.get('first_name', user.first_name).strip()
            user.last_name  = request.POST.get('last_name', user.last_name).strip()
            email_nuevo     = request.POST.get('email', '').strip()
            if email_nuevo:
                user.email = email_nuevo
            user.save(update_fields=['first_name', 'last_name', 'email'])

            messages.success(request, 'Empleado actualizado correctamente.')

        elif accion == 'reset_password':
            nueva   = request.POST.get('nueva_password', '').strip()
            nueva2  = request.POST.get('nueva_password2', '').strip()
            if not nueva:
                messages.error(request, 'Escribe la nueva contraseña.')
            elif len(nueva) < 6:
                messages.error(request, 'La contraseña debe tener al menos 6 caracteres.')
            elif nueva != nueva2:
                messages.error(request, 'Las contraseñas no coinciden.')
            else:
                empleado.perfil.user.set_password(nueva)
                empleado.perfil.user.save()
                messages.success(request, 'Contraseña actualizada correctamente.')

        return redirect('app_admin:empleado_detalle', pk=pk)

    from app_empleado.models import VentaEmpleado
    ventas_recientes = VentaEmpleado.objects.filter(empleado=empleado).order_by('-fecha')[:10]

    return render(request, 'app_admin/empleado_detalle.html', _ctx({
        'empleado':         empleado,
        'sedes':            sedes,
        'cargos':           empleado.CARGO_CHOICES,
        'ventas_recientes': ventas_recientes,
    }))


@admin_required
def empleado_toggle(request, pk):
    from authentication.models import Empleado

    empleado = get_object_or_404(Empleado, pk=pk)
    perfil   = empleado.perfil
    perfil.activo = not perfil.activo
    perfil.save(update_fields=['activo'])
    estado = 'activado' if perfil.activo else 'desactivado'
    messages.success(request, f'Empleado {estado} correctamente.')
    return redirect('app_admin:empleados')


# ── SEDES ──────────────────────────────────────────────────────────────────────

@admin_required
def sedes(request):
    from authentication.models import Sede

    if request.method == 'POST':
        nombre    = request.POST.get('nombre', '').strip()
        direccion = request.POST.get('direccion', '').strip()
        ciudad    = request.POST.get('ciudad', 'Cali').strip()

        if not nombre:
            messages.error(request, 'El nombre de la sede es obligatorio.')
        elif Sede.objects.filter(nombre__iexact=nombre).exists():
            messages.error(request, f'Ya existe una sede con el nombre "{nombre}".')
        else:
            Sede.objects.create(nombre=nombre, direccion=direccion, ciudad=ciudad)
            messages.success(request, f'Sede "{nombre}" creada correctamente.')
        return redirect('app_admin:sedes')

    qs = Sede.objects.annotate(num_empleados=Count('empleados')).order_by('nombre')
    return render(request, 'app_admin/sedes.html', _ctx({'sedes': qs}))


@admin_required
def sede_toggle(request, pk):
    from authentication.models import Sede

    sede = get_object_or_404(Sede, pk=pk)
    sede.activa = not sede.activa
    sede.save(update_fields=['activa'])
    estado = 'activada' if sede.activa else 'desactivada'
    messages.success(request, f'Sede "{sede.nombre}" {estado}.')
    return redirect('app_admin:sedes')